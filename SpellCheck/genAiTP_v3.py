import graphviz
from openpyxl import load_workbook
from PyPDF2 import PdfFileMerger, PdfFileReader, PdfFileWriter
import os.path
import sys

WIDGETNAME = "widgetName"
VALUE = "value"
XPATH = "xpath"
ACTION = "action"
XPOS = "position_X"
YPOS = "position_Y"
VALIDATE = "validate"
ACTIVE = "active"
EXECUTE = "execute"
PAGEURL = "pageURL"
CATEGORY = "category"

TAGNAME = "tagName"
CVTYPE = "type"
CVPROP = "properties"
CVRESULT = "result"
CVMATCH = "match"
CVACTION = "action"

# Requests
REQUEST_HEADER_KEYS = [PAGEURL, WIDGETNAME, VALUE, XPATH, ACTION, CATEGORY, XPOS, YPOS, VALIDATE]

REQUEST_HEADER_PAGEURL_NO = REQUEST_HEADER_KEYS.index(PAGEURL) + 1

REQUEST_HEADER_WIDGETNAME_NO = REQUEST_HEADER_KEYS.index(WIDGETNAME) + 1

REQUEST_HEADER_VALUE_NO = REQUEST_HEADER_KEYS.index(VALUE) + 1

REQUEST_HEADER_XPATH_NO = REQUEST_HEADER_KEYS.index(XPATH) + 1

REQUEST_HEADER_ACTION_NO = REQUEST_HEADER_KEYS.index(ACTION) + 1

REQUEST_HEADER_CATEGORY_NO = REQUEST_HEADER_KEYS.index(CATEGORY) + 1

REQUEST_HEADER_XPOS_NO = REQUEST_HEADER_KEYS.index(XPOS) + 1

REQUEST_HEADER_YPOS_NO = REQUEST_HEADER_KEYS.index(YPOS) + 1

REQUEST_HEADER_VALIDATE_NO = REQUEST_HEADER_KEYS.index(VALIDATE) + 1

# Response
RESPONSE_HEADER_KEYS = [PAGEURL, WIDGETNAME, VALUE, XPATH, ACTION, CATEGORY, XPOS, YPOS, ACTIVE, EXECUTE]

RESPONSE_HEADER_PAGEURL_NO = RESPONSE_HEADER_KEYS.index(PAGEURL) + 1

RESPONSE_HEADER_WIDGETNAME_NO = RESPONSE_HEADER_KEYS.index(WIDGETNAME) + 1

RESPONSE_HEADER_VALUE_NO = RESPONSE_HEADER_KEYS.index(VALUE) + 1

RESPONSE_HEADER_XPATH_NO = RESPONSE_HEADER_KEYS.index(XPATH) + 1

RESPONSE_HEADER_ACTION_NO = RESPONSE_HEADER_KEYS.index(ACTION) + 1

RESPONSE_HEADER_CATEGORY_NO = RESPONSE_HEADER_KEYS.index(CATEGORY) + 1

RESPONSE_HEADER_XPOS_NO = RESPONSE_HEADER_KEYS.index(XPOS) + 1

RESPONSE_HEADER_YPOS_NO = RESPONSE_HEADER_KEYS.index(YPOS) + 1

RESPONSE_HEADER_ACTIVE_NO = RESPONSE_HEADER_KEYS.index(ACTIVE) + 1

RESPONSE_HEADER_EXECUTE_NO = RESPONSE_HEADER_KEYS.index(EXECUTE) + 1

req_headers_length = len(REQUEST_HEADER_KEYS)
resp_headers_length = len(RESPONSE_HEADER_KEYS)
no_of_extra_rows = 7
stepOneRowNumber = 10
extraCols = 2  # Accounts for extra columns A & B (Request Items Count Value)


class GraphNodes:

    def __init__(self, excelDir, graphName, outDir, serverURL, labelNum):
        self.workbook = load_workbook(excelDir, data_only=True)
        self.totalExcelSheetNumber = len(self.workbook.sheetnames)
        self.graphName = graphName
        self.outDir = outDir
        self.serverURL = serverURL
        self.nonActionItemCondition = labelNum
        self.pageUrlRequest = []
        self.widgetNameRequest = []
        self.valueRequest = []
        self.xpathRequest = []
        self.actionRequest = []
        self.categoryRequest = []
        self.xposRequest = []
        self.yposRequest = []
        self.validateRequest = []
        self.requestDictionary = {'PageUrl': self.pageUrlRequest, 'WidgetName': self.widgetNameRequest,
                                  'Value': self.valueRequest, 'XPATH': self.xpathRequest,
                                  'Action': self.actionRequest, 'Category': self.categoryRequest,
                                  'XPOS': self.xposRequest, 'YPOS': self.yposRequest,
                                  'Validate': self.validateRequest}
        self.pageUrlResponse = []
        self.widgetNameResponse = []
        self.valueResponse = []
        self.xpathResponse = []
        self.actionResponse = []
        self.categoryResponse = []
        self.xposResponse = []
        self.yposResponse = []
        self.activeResponse = []
        self.executeResponse = []
        self.responseDictionary = {'PageUrl': self.pageUrlResponse, 'WidgetName': self.widgetNameResponse,
                                   'Value': self.valueResponse, 'XPATH': self.xpathResponse,
                                   'Action': self.actionResponse, 'Category': self.categoryResponse,
                                   'XPOS': self.xposResponse, 'YPOS': self.yposResponse,
                                   'Active': self.activeResponse, 'Execute': self.executeResponse}
        self.mainNodeArray = []
        self.subNodeArray = []
        self.subFinder = 0
        self.previousNodeArray = []
        self.currentLabelArray = []
        self.mainLabelArray = []
        self.mainNodeList = []
        self.widgetValue = "0"
        self.color = "0"
        self.shape = "0"
        self.label = "0"
        self.mainEdgeDict = {}
        self.mainEdgeDictL = {}
        self.subEdgeDict = {}
        self.xpathValue = {"Widget Value": self.widgetValue, "Color": self.color}
        self.mainNodeDict = {}
        self.mainNodeDictL = {}
        self.subNodeDict = {}
        self.edgeArray = []
        self.graphDict = {"Main Nodes": self.mainNodeArray, "Sub Nodes": self.subNodeArray, "Edges": self.edgeArray}
        self.headerSheetName = []
        self.headerPreReq = []
        self.headerStepNum = []
        self.headerPageNum = []
        self.headerPart = []
        self.htmlHeaderPart = []

    def clearArray(self):
        self.pageUrlRequest.clear()
        self.widgetNameRequest.clear()
        self.valueRequest.clear()
        self.xpathRequest.clear()
        self.actionRequest.clear()
        self.categoryRequest.clear()
        self.xposRequest.clear()
        self.yposRequest.clear()
        self.validateRequest.clear()
        self.pageUrlResponse.clear()
        self.widgetNameResponse.clear()
        self.valueResponse.clear()
        self.xpathResponse.clear()
        self.actionResponse.clear()
        self.categoryResponse.clear()
        self.xposResponse.clear()
        self.yposResponse.clear()
        self.activeResponse.clear()
        self.executeResponse.clear()

    def clearDict(self):
        self.mainNodeDict.clear()
        self.mainEdgeDict.clear()
        self.subNodeDict.clear()
        self.subEdgeDict.clear()

    def sort(self):
        mergedPdf = PdfFileMerger()
        for sheetIndex in range(self.totalExcelSheetNumber):
            if sheetIndex > 0:
                self.clearArray()
            ws = self.workbook[self.workbook.sheetnames[sheetIndex]]
            noOfStepValue = ws.cell(no_of_extra_rows + 1, 2).value
            noOfStep = int(noOfStepValue.split('=')[1])
            numberOfDataset = 1
            label = "A"
            if sheetIndex == 0:
                for header in range(self.totalExcelSheetNumber):
                    ws_header = self.workbook[self.workbook.sheetnames[header]]
                    version = ws_header.cell(5, 2).value
                    software = ws_header.cell(5, 3).value
                    createdOn = ws_header.cell(6, 2).value
                    self.headerSheetName.append("Sheet Name: " + str(ws_header.cell(1, 2).value))
                    self.headerPreReq.append("Pre Req: " + str(ws_header.cell(7, 1).value))
                    self.headerStepNum.append("Number of Steps: " + str(ws_header.cell(8, 2).value).split('=')[1])
                    self.headerPageNum.append("Page Number: " + str(header + 2))
                    self.headerPart.append(self.headerSheetName[header] + " " + self.headerPreReq[header] + " " +
                                           self.headerStepNum[header] + " " + self.headerPageNum[header])
                versionLabel = "Version"
                startInfo = f"{version} \n {software} \n"
                createdLabel = f"Created On:"
                createdInfo = createdOn
                legend1 = f"Legend:"
                legend2 = f"Pages start and end with a URL link"
                legend3 = f"Green boxes are actionable items"
                legend4 = f"Orange boxes are items that require you to fill in data"
                legend5 = f"Red regions are request blocks that follow a specific sequence to give a response"
                # fullLogoAddress = "C:/Users/Shaya/PycharmProjects/AiTestPro-TestflowGraph/pyinstaller/pythonExe/aiTestProPictures/fullLogo.png"
                dirName = os.path.dirname(__file__)
                print(f"Dirname: {dirName}")

                # determine if application is a script file or frozen exe
                if getattr(sys, 'frozen', False):
                    application_path = os.path.dirname(sys.executable)
                elif __file__:
                    application_path = os.path.dirname(__file__)
                print(f"Application path: {application_path}")
                fileName = os.path.join(application_path, 'images', 'dash.png')

                print(f"File name : {fileName}")

                fullLogoAddress = fileName
                border = '0'
                cellBorder = '1'
                width = '1'
                height = '1'
                tableStart = f"<<table border='{border}' cellborder='{cellBorder}' width='{width}' height='{height}'>"
                tableEnd = f"</table>>"
                color = 'black'
                legendBody1 = f"<tr><td ALIGN='LEFT'><FONT COLOR='{color}'>{legend1}</FONT></td></tr>"
                legendBody2 = f"<tr><td ALIGN='LEFT'><FONT COLOR='{color}'>{legend2}</FONT></td></tr>"
                legendBody3 = f"<tr><td ALIGN='LEFT'><FONT COLOR='{color}'>{legend3}</FONT></td></tr>"
                legendBody4 = f"<tr><td ALIGN='LEFT'><FONT COLOR='{color}'>{legend4}</FONT></td></tr>"
                legendBody5 = f"<tr><td ALIGN='LEFT'><FONT COLOR='{color}'>{legend5}</FONT></td></tr>"
                bodyFullImage = f"<tr><td colspan='4'><IMG src='{fullLogoAddress}'/></td></tr>"
                verBody = f"<tr><td ALIGN='LEFT'><FONT COLOR='{color}'>{versionLabel}</FONT></td><td>{startInfo}</td></tr>"
                createdBody = f"<tr><td ALIGN='LEFT'><FONT COLOR='{color}'>{createdLabel}</FONT></td><td>{createdInfo}</td></tr>"
                bodyDefaultInfo = f"<tr><td ALIGN='LEFT'><FONT COLOR='{color}'>{startInfo}</FONT></td></tr>"
                finalLabel = tableStart + bodyFullImage + verBody + createdBody
                for body in range(len(self.headerPart)):
                    self.htmlHeaderPart.append(f"<tr><td ALIGN='LEFT'><FONT COLOR='{color}'>{self.headerPart[body]}</FONT></td></tr>")
                    finalLabel += self.htmlHeaderPart[body]
                finalLabel += legendBody1
                finalLabel += legendBody2
                finalLabel += legendBody3
                finalLabel += legendBody4
                finalLabel += legendBody5
                finalLabel += tableEnd
                self.graphHeader(mergedPdf, finalLabel)

            for dataNum in range(numberOfDataset):
                datasetHeaderRowNum = int(no_of_extra_rows) + 1 + ((2 * int(noOfStep) + 1) * int(dataNum))
                for step in range(noOfStep):
                    edgeDict = {}
                    # if sheetIndex > 0:
                    #     step += 1
                    stepHeadersRow = datasetHeaderRowNum + 1 + (2 * int(step))
                    if stepHeadersRow >= ws.max_row:
                        break
                    stepValuesRow = stepHeadersRow + 1
                    requestItemsCount = int(ws.cell(stepValuesRow, 2).value)
                    requestItemsCell = 2
                    # Start at requestItemsCell, check the request items, then the following +1 cell is the response
                    responseItemsCell = requestItemsCell + (requestItemsCount * req_headers_length) + 1
                    responseItemsCount = int(ws.cell(stepValuesRow, responseItemsCell).value)
                    self.mainNodeArray.append(ws.cell(stepOneRowNumber,
                                                      REQUEST_HEADER_VALUE_NO + extraCols).value + " " +
                                              ws.cell(stepOneRowNumber,
                                                      REQUEST_HEADER_ACTION_NO + extraCols).value)
                    openUrlXpath = ws.cell(stepOneRowNumber,
                                           REQUEST_HEADER_XPATH_NO + extraCols).value
                    if openUrlXpath is None or " ":
                        openUrlXpath = ws.cell(stepOneRowNumber,
                                               REQUEST_HEADER_VALUE_NO + extraCols).value
                    firstDict = {}
                    firstDict["Widget"] = ws.cell(stepOneRowNumber,
                                                  REQUEST_HEADER_VALUE_NO + extraCols).value + " " + ws.cell(
                        stepOneRowNumber,
                        REQUEST_HEADER_ACTION_NO + extraCols).value
                    previousLabel = "A0"
                    firstDict["Label"] = previousLabel
                    firstDict["color"] = "black"
                    firstDict["shape"] = "note"
                    self.mainNodeDict[openUrlXpath] = firstDict
                    nodeCounterResponse = 1
                    nodeCounterRequest = 1
                    nodeCounter = 1
                    if step > 0:
                        requestXpath = ws.cell(stepValuesRow, requestItemsCell + (req_headers_length * (
                                requestItemsCount - 1)) + REQUEST_HEADER_XPATH_NO).value
                        for reqPull in range(requestItemsCount):
                            self.pageUrlRequest.append(ws.cell(stepValuesRow, requestItemsCell + (
                                    req_headers_length * reqPull) + REQUEST_HEADER_PAGEURL_NO).value)
                            self.widgetNameRequest.append(ws.cell(stepValuesRow, requestItemsCell + (
                                    req_headers_length * reqPull) + REQUEST_HEADER_WIDGETNAME_NO).value)
                            self.valueRequest.append(ws.cell(stepValuesRow, requestItemsCell + (
                                    req_headers_length * reqPull) + REQUEST_HEADER_VALUE_NO).value)
                            self.xpathRequest.append(ws.cell(stepValuesRow, requestItemsCell + (
                                    req_headers_length * reqPull) + REQUEST_HEADER_XPATH_NO).value)
                            self.actionRequest.append(ws.cell(stepValuesRow, requestItemsCell + (
                                    req_headers_length * reqPull) + REQUEST_HEADER_ACTION_NO).value)
                            self.categoryRequest.append(ws.cell(stepValuesRow, requestItemsCell + (
                                    req_headers_length * reqPull) + REQUEST_HEADER_CATEGORY_NO).value)
                            if reqPull < requestItemsCount - 1:
                                if self.categoryRequest[reqPull] == "MenuLink":
                                    self.shape = "cds"
                                    self.color = "green"
                                elif self.categoryRequest[reqPull] == "FormElement":
                                    self.shape = "note"
                                    self.color = "orange"
                                elif self.categoryRequest[reqPull] == "LabelsText":
                                    self.shape = "invhouse"
                                    self.color = "black"
                                else:
                                    self.shape = "underline"
                                    self.color = "grey11"
                                simpleWidget = self.widgetNameRequest[reqPull].strip().replace('\\', '')
                                simpleValue = self.valueRequest[reqPull].strip().replace('\\', '')
                                self.subNodeArray.append(self.widgetNameRequest[reqPull])
                                if self.mainNodeDict.get(self.xpathRequest[reqPull] + self.pageUrlRequest[reqPull]):
                                    tempLabel = self.mainNodeDict.get(
                                        self.xpathRequest[reqPull] + self.pageUrlRequest[reqPull]).get('Label')
                                    keyListEdge = list(self.mainEdgeDict.keys())
                                    for indexEdge in range(len(keyListEdge)):
                                        mainEdge = keyListEdge[indexEdge]
                                        if tempLabel in mainEdge:
                                            self.mainEdgeDict.pop(mainEdge)
                                    self.mainNodeDict.pop(self.xpathRequest[reqPull] + self.pageUrlRequest[reqPull])
                                    firstDict = {}
                                    firstDict["Widget"] = simpleWidget  # + " " + simpleValue
                                    firstDict["Label"] = label + str(nodeCounter)
                                    firstDict["color"] = self.color
                                    firstDict["shape"] = self.shape
                                    self.subNodeDict[
                                        self.xpathRequest[reqPull] + self.pageUrlRequest[reqPull]] = firstDict
                                    edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                    if reqPull == 0:
                                        self.mainEdgeDict[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                    if reqPull > 0:
                                        self.subEdgeDict[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                    previousLabel = label + str(nodeCounter)
                                    nodeCounter += 1
                        try:
                            xpathIndex = self.xpathRequest.index(requestXpath)
                            self.subNodeArray.append(self.widgetNameRequest[xpathIndex])
                            if self.mainNodeDict.get(self.xpathRequest[reqPull] + self.pageUrlRequest[reqPull]):
                                tempLabel = self.mainNodeDict.get(
                                    self.xpathRequest[reqPull] + self.pageUrlRequest[reqPull]).get('Label')
                                keyListEdge = list(self.mainEdgeDict.keys())
                                for indexEdge in range(len(keyListEdge)):
                                    mainEdge = keyListEdge[indexEdge]
                                    if tempLabel in mainEdge:
                                        self.mainEdgeDict.pop(mainEdge)
                                self.mainNodeDict.pop(self.xpathRequest[reqPull] + self.pageUrlRequest[reqPull])
                                self.subFinder = reqPull
                                if self.categoryRequest[reqPull] == "MenuLink":
                                    self.shape = "cds"
                                    self.color = "green"
                                elif self.categoryRequest[reqPull] == "FormElement":
                                    self.shape = "note"
                                    self.color = "orange"
                                elif self.categoryRequest[reqPull] == "LabelsText":
                                    self.shape = "invhouse"
                                    self.color = "black"
                                else:
                                    self.shape = "underline"
                                    self.color = "grey11"
                                simpleWidget = self.widgetNameRequest[reqPull].strip().replace('\\', '')
                                simpleValue = self.valueRequest[reqPull].strip().replace('\\', '')
                                self.subNodeArray.append(self.widgetNameRequest[reqPull])
                                firstDict = {}
                                firstDict["Widget"] = simpleWidget + " " + simpleValue
                                firstDict["Label"] = label + str(nodeCounter)
                                firstDict["color"] = self.color
                                firstDict["shape"] = self.shape
                                self.subNodeDict[self.xpathRequest[reqPull] + self.pageUrlRequest[reqPull]] = firstDict
                                edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                self.subEdgeDict[edgeKey] = [previousLabel, label + str(nodeCounter)]
                        except ValueError:
                            break
                        previousLabel = label + str(nodeCounter)
                        nodeCounter += 1
                    self.clearArray()
                    for respPull in range(responseItemsCount):
                        cellValueTemp = requestItemsCell + (requestItemsCount * req_headers_length) + (
                                resp_headers_length * respPull) + 1
                        if respPull == 0 and step > 0:
                            newLabel = previousLabel
                            openUrlXpath = ws.cell(stepValuesRow,
                                                   cellValueTemp + RESPONSE_HEADER_PAGEURL_NO).value
                            firstDict = {}
                            firstDict["Widget"] = ws.cell(stepValuesRow,
                                                          cellValueTemp + RESPONSE_HEADER_PAGEURL_NO).value
                            newLabel = label + str(nodeCounter)
                            firstDict["Label"] = newLabel
                            firstDict["color"] = "black"
                            firstDict["shape"] = "note"
                            self.mainNodeDict[openUrlXpath] = firstDict
                            edgeKeyNewPage = previousLabel + "_" + newLabel
                            self.mainEdgeDict[edgeKeyNewPage] = [previousLabel, newLabel]
                            nodeCounter += 1
                            previousLabel = newLabel
                            firstDict = {}
                            break
                        self.pageUrlResponse.append(
                            ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_PAGEURL_NO).value)
                        self.widgetNameResponse.append(
                            ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_WIDGETNAME_NO).value)
                        self.valueResponse.append(
                            ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_VALUE_NO).value)
                        self.xpathResponse.append(
                            ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_XPATH_NO).value)
                        self.actionResponse.append(
                            ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_ACTION_NO).value)
                        self.activeResponse.append(
                            ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_ACTIVE_NO).value)
                        self.executeResponse.append(
                            ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_EXECUTE_NO).value)
                        self.categoryResponse.append(
                            ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_CATEGORY_NO).value)
                        if str(self.executeResponse[respPull]).lower() == "true":
                            if self.nonActionItemCondition == 1:
                                if str(self.activeResponse[respPull]).lower() == "true":
                                    simpleWidget = self.widgetNameResponse[respPull].strip().replace('\\', '')
                                    simpleValue = self.valueResponse[respPull].strip().replace('\\', '')
                                    if self.categoryResponse[respPull] == "MenuLink":
                                        self.shape = "cds"
                                        self.color = "green"
                                    elif self.categoryResponse[respPull] == "FormElement":
                                        self.shape = "note"
                                        self.color = "orange"
                                    elif self.categoryResponse[respPull] == "LabelsText":
                                        self.shape = "invhouse"
                                        self.color = "black"
                                    else:
                                        self.shape = "underline"
                                        self.color = "grey11"
                                    if self.mainNodeDict.get(
                                            self.xpathResponse[respPull] + self.pageUrlResponse[
                                                respPull]) is None:
                                        if self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]) is None:
                                            firstDict = {}
                                            if simpleWidget == simpleValue:
                                                self.mainNodeArray.append(self.valueResponse[respPull])
                                                firstDict["Widget"] = self.valueResponse[respPull]
                                            elif simpleValue:
                                                self.mainNodeArray.append(
                                                    self.valueResponse[respPull] + " " + self.widgetNameResponse[
                                                        respPull])
                                                firstDict["Widget"] = self.valueResponse[respPull] + " " + \
                                                                      self.widgetNameResponse[
                                                                          respPull]
                                            else:
                                                self.mainNodeArray.append(self.widgetNameResponse[respPull])
                                                firstDict["Widget"] = self.widgetNameResponse[respPull]
                                            firstDict["Label"] = label + str(nodeCounter)
                                            firstDict["color"] = self.color
                                            firstDict["shape"] = self.shape
                                            self.mainNodeDict[
                                                self.xpathResponse[respPull] + self.pageUrlResponse[
                                                    respPull]] = firstDict
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            self.mainEdgeDict[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                            nodeCounter += 1
                                        else:
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            edgeKey2 = previousLabel + "_" + self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                'Label')
                                            self.subEdgeDict[edgeKey2] = [previousLabel, self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                'Label')]
                                    else:
                                        edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                        if self.mainNodeDict.get(self.xpathResponse[respPull]):
                                            edgeKey2 = previousLabel + "_" + self.mainNodeDict.get(
                                                self.xpathResponse[respPull]).get('Label')
                                            self.mainEdgeDict[edgeKey2] = [previousLabel, self.mainNodeDict.get(
                                                self.xpathResponse[respPull]).get('Label')]
                            elif self.nonActionItemCondition == 2:
                                if str(self.activeResponse[respPull]).lower() != "true":
                                    self.shape = "box"
                                    self.color = "grey11"
                                    simpleWidget = self.widgetNameResponse[respPull].strip().replace('\\', '')
                                    simpleValue = self.valueResponse[respPull].strip().replace('\\', '')
                                    if simpleWidget == simpleValue:
                                        widgetValue = simpleValue
                                    elif simpleValue:
                                        widgetValue = simpleValue + " " + simpleWidget
                                    else:
                                        widgetValue = simpleWidget
                                    if self.mainNodeDict.get(
                                            "sheet" + str(sheetIndex) + "_step" + str(step) + "_labels") is None:
                                        firstDict = {}
                                        firstDict["Widget"] = "Labels: " + "\n \n" + widgetValue
                                        firstDict["Label"] = label + str(nodeCounter)
                                        firstDict["color"] = self.color
                                        firstDict["shape"] = self.shape
                                        self.mainNodeDict[
                                            "sheet" + str(sheetIndex) + "_step" + str(step) + "_labels"] = firstDict
                                        edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                        self.mainEdgeDict[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                        nodeCounter += 1
                                    else:
                                        self.mainNodeDict["sheet" + str(sheetIndex) + "_step" + str(step) + "_labels"][
                                            "Widget"] = self.mainNodeDict.get(
                                            "sheet" + str(sheetIndex) + "_step" + str(step) + "_labels").get(
                                            "Widget") + "\n" + widgetValue
                                else:
                                    simpleWidget = self.widgetNameResponse[respPull].strip().replace('\\', '')
                                    simpleValue = self.valueResponse[respPull].strip().replace('\\', '')
                                    if self.categoryResponse[respPull] == "MenuLink":
                                        self.shape = "cds"
                                        self.color = "green"
                                    elif self.categoryResponse[respPull] == "FormElement":
                                        self.shape = "note"
                                        self.color = "orange"
                                    elif self.categoryResponse[respPull] == "LabelsText":
                                        self.shape = "invhouse"
                                        self.color = "black"
                                    else:
                                        self.shape = "underline"
                                        self.color = "grey11"
                                    if self.mainNodeDict.get(
                                            self.xpathResponse[respPull] + self.pageUrlResponse[
                                                respPull]) is None:  # or " ":
                                        if self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]) is None:
                                            firstDict = {}
                                            if simpleWidget == simpleValue:
                                                self.mainNodeArray.append(self.valueResponse[respPull])
                                                firstDict["Widget"] = self.valueResponse[respPull]
                                            elif simpleValue:
                                                self.mainNodeArray.append(
                                                    self.valueResponse[respPull] + " " + self.widgetNameResponse[
                                                        respPull])
                                                firstDict["Widget"] = self.valueResponse[respPull] + " " + \
                                                                      self.widgetNameResponse[
                                                                          respPull]
                                            else:
                                                self.mainNodeArray.append(self.widgetNameResponse[respPull])
                                                firstDict["Widget"] = self.widgetNameResponse[respPull]
                                            firstDict["Label"] = label + str(nodeCounter)
                                            firstDict["color"] = self.color
                                            firstDict["shape"] = self.shape
                                            self.mainNodeDict[
                                                self.xpathResponse[respPull] + self.pageUrlResponse[
                                                    respPull]] = firstDict
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            self.mainEdgeDict[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                            nodeCounter += 1
                                        else:
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            edgeKey2 = previousLabel + "_" + self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                'Label')
                                            self.subEdgeDict[edgeKey2] = [previousLabel, self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                'Label')]
                                    else:
                                        edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                        if self.mainNodeDict.get(self.xpathResponse[respPull]):
                                            edgeKey2 = previousLabel + "_" + self.mainNodeDict.get(
                                                self.xpathResponse[respPull]).get('Label')
                                            self.mainEdgeDict[edgeKey2] = [previousLabel, self.mainNodeDict.get(
                                                self.xpathResponse[respPull]).get('Label')]
                            else:
                                simpleWidget = self.widgetNameResponse[respPull].strip().replace('\\', '')
                                simpleValue = self.valueResponse[respPull].strip().replace('\\', '')
                                if self.categoryResponse[respPull] == "MenuLink":
                                    self.shape = "cds"
                                    self.color = "green"
                                elif self.categoryResponse[respPull] == "FormElement":
                                    self.shape = "note"
                                    self.color = "orange"
                                elif self.categoryResponse[respPull] == "LabelsText":
                                    self.shape = "invhouse"
                                    self.color = "black"
                                else:
                                    self.shape = "underline"
                                    self.color = "grey11"
                                if self.mainNodeDict.get(
                                        self.xpathResponse[respPull] + self.pageUrlResponse[
                                            respPull]) is None:
                                    if self.subNodeDict.get(
                                            self.xpathResponse[respPull] + self.pageUrlResponse[respPull]) is None:
                                        firstDict = {}
                                        if simpleWidget == simpleValue:
                                            self.mainNodeArray.append(self.valueResponse[respPull])
                                            firstDict["Widget"] = self.valueResponse[respPull]
                                        elif simpleValue:
                                            self.mainNodeArray.append(
                                                self.valueResponse[respPull] + " " + self.widgetNameResponse[respPull])
                                            firstDict["Widget"] = self.valueResponse[respPull] + " " + \
                                                                  self.widgetNameResponse[
                                                                      respPull]
                                        else:
                                            self.mainNodeArray.append(self.widgetNameResponse[respPull])
                                            firstDict["Widget"] = self.widgetNameResponse[respPull]
                                        firstDict["Label"] = label + str(nodeCounter)
                                        firstDict["color"] = self.color
                                        firstDict["shape"] = self.shape
                                        self.mainNodeDict[
                                            self.xpathResponse[respPull] + self.pageUrlResponse[respPull]] = firstDict
                                        edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                        self.mainEdgeDict[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                        nodeCounter += 1
                                    else:
                                        edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                        edgeKey2 = previousLabel + "_" + self.subNodeDict.get(
                                            self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get('Label')
                                        self.subEdgeDict[edgeKey2] = [previousLabel, self.subNodeDict.get(
                                            self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get('Label')]
                                else:
                                    edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                    if self.mainNodeDict.get(self.xpathResponse[respPull]):
                                        edgeKey2 = previousLabel + "_" + self.mainNodeDict.get(
                                            self.xpathResponse[respPull]).get('Label')
                                        self.mainEdgeDict[edgeKey2] = [previousLabel, self.mainNodeDict.get(
                                            self.xpathResponse[respPull]).get('Label')]
                    ###################################################################
                    if sheetIndex == (self.totalExcelSheetNumber - 1) and step == (noOfStep - 1):
                        for respPull in range(responseItemsCount):
                            cellValueTemp = requestItemsCell + (requestItemsCount * req_headers_length) + (
                                    resp_headers_length * respPull) + 1
                            if respPull == 0 and step > 0:
                                newLabel = previousLabel
                                openUrlXpath = ws.cell(stepValuesRow,
                                                       cellValueTemp + RESPONSE_HEADER_PAGEURL_NO).value
                                firstDict = {}
                                firstDict["Widget"] = ws.cell(stepValuesRow,
                                                              cellValueTemp + RESPONSE_HEADER_PAGEURL_NO).value
                                newLabel = label + str(nodeCounter)
                                firstDict["Label"] = newLabel
                                firstDict["color"] = "black"
                                firstDict["shape"] = "note"
                                self.mainNodeDictL[openUrlXpath] = firstDict
                                edgeKeyNewPage = previousLabel + "_" + newLabel
                                nodeCounter += 1
                                previousLabel = newLabel
                                firstDict = {}
                            self.pageUrlResponse.append(
                                ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_PAGEURL_NO).value)
                            self.widgetNameResponse.append(
                                ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_WIDGETNAME_NO).value)
                            self.valueResponse.append(
                                ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_VALUE_NO).value)
                            self.xpathResponse.append(
                                ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_XPATH_NO).value)
                            self.actionResponse.append(
                                ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_ACTION_NO).value)
                            self.activeResponse.append(
                                ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_ACTIVE_NO).value)
                            self.executeResponse.append(
                                ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_EXECUTE_NO).value)
                            self.categoryResponse.append(
                                ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_CATEGORY_NO).value)
                            if str(self.executeResponse[respPull]).lower() == "true":
                                if self.nonActionItemCondition == 1:
                                    if str(self.activeResponse[respPull]).lower() == "true":
                                        simpleWidget = self.widgetNameResponse[respPull].strip().replace('\\', '')
                                        simpleValue = self.valueResponse[respPull].strip().replace('\\', '')
                                        if self.categoryResponse[respPull] == "MenuLink":
                                            self.shape = "cds"
                                            self.color = "green"
                                        elif self.categoryResponse[respPull] == "FormElement":
                                            self.shape = "note"
                                            self.color = "orange"
                                        elif self.categoryResponse[respPull] == "LabelsText":
                                            self.shape = "invhouse"
                                            self.color = "black"
                                        else:
                                            self.shape = "underline"
                                            self.color = "grey11"
                                        if self.mainNodeDictL.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[
                                                    respPull]) is None:  # or " ":
                                            if self.subNodeDict.get(
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[
                                                        respPull]) is None:
                                                firstDict = {}
                                                if simpleWidget == simpleValue:
                                                    self.mainNodeArray.append(self.valueResponse[respPull])
                                                    firstDict["Widget"] = self.valueResponse[respPull]
                                                elif simpleValue:
                                                    self.mainNodeArray.append(
                                                        self.valueResponse[respPull] + " " + self.widgetNameResponse[
                                                            respPull])
                                                    firstDict["Widget"] = self.valueResponse[respPull] + " " + \
                                                                          self.widgetNameResponse[
                                                                              respPull]
                                                else:
                                                    self.mainNodeArray.append(self.widgetNameResponse[respPull])
                                                    firstDict["Widget"] = self.widgetNameResponse[respPull]
                                                firstDict["Label"] = label + str(nodeCounter)
                                                firstDict["color"] = self.color
                                                firstDict["shape"] = self.shape
                                                self.mainNodeDictL[
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[
                                                        respPull]] = firstDict
                                                edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                                self.mainEdgeDictL[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                                nodeCounter += 1
                                            else:
                                                edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                                edgeKey2 = previousLabel + "_" + self.subNodeDict.get(
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                    'Label')
                                                self.subEdgeDict[edgeKey2] = [previousLabel, self.subNodeDict.get(
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                    'Label')]
                                        else:
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            if self.mainNodeDictL.get(self.xpathResponse[respPull]):
                                                edgeKey2 = previousLabel + "_" + self.mainNodeDict.get(
                                                    self.xpathResponse[respPull]).get('Label')
                                                self.mainEdgeDictL[edgeKey2] = [previousLabel, self.mainNodeDict.get(
                                                    self.xpathResponse[respPull]).get('Label')]
                                elif self.nonActionItemCondition == 2:
                                    if str(self.activeResponse[respPull]).lower() != "true":
                                        self.shape = "box"
                                        self.color = "grey11"
                                        simpleWidget = self.widgetNameResponse[respPull].strip().replace('\\', '')
                                        simpleValue = self.valueResponse[respPull].strip().replace('\\', '')
                                        if simpleWidget == simpleValue:
                                            widgetValue = simpleValue
                                        elif simpleValue:
                                            widgetValue = simpleValue + " " + simpleWidget
                                        else:
                                            widgetValue = simpleWidget
                                        if self.mainNodeDictL.get(
                                                "sheet" + str(sheetIndex) + "_step" + str(step) + "_labels") is None:
                                            firstDict = {}
                                            firstDict["Widget"] = "Labels: " + "\n \n" + widgetValue
                                            firstDict["Label"] = label + str(nodeCounter)
                                            firstDict["color"] = self.color
                                            firstDict["shape"] = self.shape
                                            self.mainNodeDictL[
                                                "sheet" + str(sheetIndex) + "_step" + str(step) + "_labels"] = firstDict
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            self.mainEdgeDictL[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                            nodeCounter += 1
                                        else:
                                            self.mainNodeDictL[
                                                "sheet" + str(sheetIndex) + "_step" + str(step) + "_labels"][
                                                "Widget"] = self.mainNodeDictL.get(
                                                "sheet" + str(sheetIndex) + "_step" + str(step) + "_labels").get(
                                                "Widget") + "\n" + widgetValue
                                    else:
                                        simpleWidget = self.widgetNameResponse[respPull].strip().replace('\\', '')
                                        simpleValue = self.valueResponse[respPull].strip().replace('\\', '')
                                        if self.categoryResponse[respPull] == "MenuLink":
                                            self.shape = "cds"
                                            self.color = "green"
                                        elif self.categoryResponse[respPull] == "FormElement":
                                            self.shape = "note"
                                            self.color = "orange"
                                        elif self.categoryResponse[respPull] == "LabelsText":
                                            self.shape = "invhouse"
                                            self.color = "black"
                                        else:
                                            self.shape = "underline"
                                            self.color = "grey11"
                                        if self.mainNodeDictL.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[
                                                    respPull]) is None:  # or " ":
                                            if self.subNodeDict.get(
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[
                                                        respPull]) is None:
                                                firstDict = {}
                                                if simpleWidget == simpleValue:
                                                    # print(f"Node: {self.valueResponse[respPull]}")
                                                    self.mainNodeArray.append(self.valueResponse[respPull])
                                                    firstDict["Widget"] = self.valueResponse[respPull]
                                                elif simpleValue:
                                                    self.mainNodeArray.append(
                                                        self.valueResponse[respPull] + " " + self.widgetNameResponse[
                                                            respPull])
                                                    firstDict["Widget"] = self.valueResponse[respPull] + " " + \
                                                                          self.widgetNameResponse[
                                                                              respPull]
                                                else:
                                                    self.mainNodeArray.append(self.widgetNameResponse[respPull])
                                                    firstDict["Widget"] = self.widgetNameResponse[respPull]
                                                firstDict["Label"] = label + str(nodeCounter)
                                                firstDict["color"] = self.color
                                                firstDict["shape"] = self.shape
                                                self.mainNodeDictL[
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[
                                                        respPull]] = firstDict
                                                edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                                self.mainEdgeDictL[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                                nodeCounter += 1
                                            else:
                                                edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                                edgeKey2 = previousLabel + "_" + self.subNodeDict.get(
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                    'Label')
                                                self.subEdgeDict[edgeKey2] = [previousLabel, self.subNodeDict.get(
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                    'Label')]
                                        else:
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            if self.mainNodeDictL.get(self.xpathResponse[respPull]):
                                                edgeKey2 = previousLabel + "_" + self.mainNodeDictL.get(
                                                    self.xpathResponse[respPull]).get('Label')
                                                self.mainEdgeDictL[edgeKey2] = [previousLabel, self.mainNodeDictL.get(
                                                    self.xpathResponse[respPull]).get('Label')]
                                else:
                                    simpleWidget = self.widgetNameResponse[respPull].strip().replace('\\', '')
                                    simpleValue = self.valueResponse[respPull].strip().replace('\\', '')
                                    if self.categoryResponse[respPull] == "MenuLink":
                                        self.shape = "cds"
                                        self.color = "green"
                                    elif self.categoryResponse[respPull] == "FormElement":
                                        self.shape = "note"
                                        self.color = "orange"
                                    elif self.categoryResponse[respPull] == "LabelsText":
                                        self.shape = "invhouse"
                                        self.color = "black"
                                    else:
                                        self.shape = "underline"
                                        self.color = "grey11"
                                    if self.mainNodeDictL.get(
                                            self.xpathResponse[respPull] + self.pageUrlResponse[
                                                respPull]) is None:  # or " ":
                                        if self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]) is None:
                                            firstDict = {}
                                            if simpleWidget == simpleValue:
                                                # print(f"Node: {self.valueResponse[respPull]}")
                                                self.mainNodeArray.append(self.valueResponse[respPull])
                                                firstDict["Widget"] = self.valueResponse[respPull]
                                            elif simpleValue:
                                                self.mainNodeArray.append(
                                                    self.valueResponse[respPull] + " " + self.widgetNameResponse[
                                                        respPull])
                                                firstDict["Widget"] = self.valueResponse[respPull] + " " + \
                                                                      self.widgetNameResponse[
                                                                          respPull]
                                            else:
                                                self.mainNodeArray.append(self.widgetNameResponse[respPull])
                                                firstDict["Widget"] = self.widgetNameResponse[respPull]
                                            firstDict["Label"] = label + str(nodeCounter)
                                            firstDict["color"] = self.color
                                            firstDict["shape"] = self.shape
                                            self.mainNodeDictL[
                                                self.xpathResponse[respPull] + self.pageUrlResponse[
                                                    respPull]] = firstDict
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            self.mainEdgeDictL[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                            nodeCounter += 1
                                        else:
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            edgeKey2 = previousLabel + "_" + self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                'Label')
                                            self.subEdgeDict[edgeKey2] = [previousLabel, self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                'Label')]
                                    else:
                                        edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                        if self.mainNodeDictL.get(self.xpathResponse[respPull]):
                                            edgeKey2 = previousLabel + "_" + self.mainNodeDict.get(
                                                self.xpathResponse[respPull]).get('Label')
                                            self.mainEdgeDictL[edgeKey2] = [previousLabel, self.mainNodeDict.get(
                                                self.xpathResponse[respPull]).get('Label')]
                    ###################################################################
                    previousLabel = label + str(nodeCounter - 1)
                    label = chr(ord(label) + 1)
            self.graph(mergedPdf, sheetIndex)
            if sheetIndex == (self.totalExcelSheetNumber - 1):
                self.graphL(mergedPdf, sheetIndex)
            self.clearDict()

        mergedPdf.write(f"{self.outDir}{self.graphName}.pdf")
        os.remove(f"{self.outDir}{self.graphName}")
        print(f"{self.serverURL}{self.graphName}.pdf")

    def graphHeader(self, mergePdf, label):
        graph = graphviz.Digraph(self.graphName)
        graph.node('A1', shape='none', color='black', label=label)
        graph.render(filename=f"{self.outDir}{self.graphName}", view=False)
        mergePdf.append(PdfFileReader(f"{self.outDir}{self.graphName}.pdf", 'rb'))
        os.remove(graph.render(filename=f"{self.outDir}{self.graphName}", view=False))

    def graph(self, mergePdf, sheetPage):
        graph = graphviz.Digraph(self.graphName)
        graph.attr(shape="circle", color="green", compound="True")

        keyList = list(self.mainNodeDict.keys())
        for index in range(len(keyList)):
            mainNode = keyList[index]
            graph.node(self.mainNodeDict.get(mainNode).get('Label'), self.mainNodeDict.get(mainNode).get('Widget'),
                       color=self.mainNodeDict.get(mainNode).get('color'),
                       shape=self.mainNodeDict.get(mainNode).get('shape'))

        keyListEdge = list(self.mainEdgeDict.keys())
        for indexEdge in range(len(keyListEdge)):
            mainEdge = keyListEdge[indexEdge]
            graph.edge(self.mainEdgeDict.get(mainEdge)[0], self.mainEdgeDict.get(mainEdge)[1])

        # Create subgraph from the main 'graph' tree
        # In order to use the cluster traits, name must start with the string "cluster"
        with graph.subgraph(name="cluster") as subCluster:
            # Compound ensures that the rank and shape of the subgraph would not
            # take precedence over the main graph node structure hierarchy
            subCluster.attr(shape="box", color="red", compound="True")
            # Convert the dictionary used to obtain node values into a list
            keyListSub = list(self.subNodeDict.keys())
            for indexSub in range(len(keyListSub)):
                subNode = keyListSub[indexSub]
                # Create node using the Label and Values derived from the dictionary
                # Dictionary also holds shape and color identity
                subCluster.node(self.subNodeDict.get(subNode).get('Label'), self.subNodeDict.get(subNode).get('Widget'),
                                color=self.subNodeDict.get(subNode).get('color'),
                                shape=self.subNodeDict.get(subNode).get('shape'))
            # Loop over the edge dictionary to make the connections between the nodes
            # Edges will be connecting using the nodes' labels
            # For example, nodes (A0, X) & (A1, Y) will be connected by edge:
            # (A0, A1), which will result in the tree creating the connection that (X -> Y) in the child hierarchy
            keyListEdgeSub = list(self.subEdgeDict.keys())
            for indexEdgeSub in range(len(keyListEdgeSub)):
                subEdge = keyListEdgeSub[indexEdgeSub]
                subCluster.edge(self.subEdgeDict.get(subEdge)[0], self.subEdgeDict.get(subEdge)[1])


        graph.render(filename=f"{self.outDir}{self.graphName}", view=False)
        mergePdf.append(PdfFileReader(f"{self.outDir}{self.graphName}.pdf", 'rb'))
        os.remove(graph.render(filename=f"{self.outDir}{self.graphName}", view=False))


    def graphL(self, mergePdf, sheetPage):
        graph = graphviz.Digraph(self.graphName)
        graph.attr(shape="circle", color="green", compound="True")

        keyList = list(self.mainNodeDictL.keys())
        for index in range(len(keyList)):
            mainNode = keyList[index]
            # print(f"Main node_graph: {mainNode}")
            # print(f"Dict main node: {self.mainNodeDict.get(mainNode)}")
            # print(f"Dict main node: {self.mainNodeDict.get(mainNode).get('Widget')}")
            graph.node(self.mainNodeDictL.get(mainNode).get('Label'), self.mainNodeDictL.get(mainNode).get('Widget'),
                       color=self.mainNodeDictL.get(mainNode).get('color'),
                       shape=self.mainNodeDictL.get(mainNode).get('shape'))

        keyListEdge = list(self.mainEdgeDictL.keys())
        for indexEdge in range(len(keyListEdge)):
            mainEdge = keyListEdge[indexEdge]
            graph.edge(self.mainEdgeDictL.get(mainEdge)[0], self.mainEdgeDictL.get(mainEdge)[1])



        graph.render(filename=f"{self.outDir}{self.graphName}", view=False)
        mergePdf.append(PdfFileReader(f"{self.outDir}{self.graphName}.pdf", 'rb'))
        os.remove(graph.render(filename=f"{self.outDir}{self.graphName}", view=False))



def main(excelDir, graphName, outDir, serverURL, labelNum):
    wb = GraphNodes(excelDir, graphName, outDir, serverURL, labelNum)
    wb.sort()


if __name__ == "__main__":
    # excelPath = sys.argv[1]
    # graphvizName = sys.argv[2]
    # outputDirectory = sys.argv[3]
    # serverUrl = sys.argv[4]
    # labelType = int(sys.argv[5])
    excelPath = "C:/Users/Shaya/PycharmProjects/AiTestPro-TestflowGraph/pyinstaller/pythonExe/fixed_DGErrorLeema28_01.xlsx"
    graphvizName = "Header_Fixed_DGErrorLeema_Label2"
    outputDirectory = "C:/Users/Shaya/PycharmProjects/AiTestPro-TestflowGraph/pyinstaller/pythonExe/dist"
    serverUrl = "testUrl"
    labelType = 2
    main(excelPath, graphvizName, outputDirectory, serverUrl, labelType)
