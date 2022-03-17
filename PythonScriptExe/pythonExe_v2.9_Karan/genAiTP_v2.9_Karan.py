import textwrap
import graphviz
from openpyxl import load_workbook
from PyPDF2 import PdfFileMerger, PdfFileReader, PdfFileWriter
from PyPDF2.pdf import PageObject
import os.path
import sys
import fitz
from PIL import Image, ImageFont, ImageDraw
import urllib.request
import simplejson as js
from fpdf import FPDF
from symspellpy import SymSpell, Verbosity
import shutil
import time

s = time.time()


WIDGETNAME = "widgetName"
VALUE = "value"
XPATH = "xpath"
ACTION = "action"
XPOS = "position_X"
YPOS = "position_Y"
VALIDATE = "validate"
ACTIVE = "active"
EXECUTE = "execute"
PAGEURL = "pageURL"
CATEGORY = "category"
MISC = "misc"
DYNAMIC = 'dynamicControl'

TAGNAME = "tagName"
CVTYPE = "type"
CVPROP = "properties"
CVRESULT = "result"
CVMATCH = "match"
CVACTION = "action"

# Requests
# REQUEST_HEADER_KEYS = [PAGEURL, WIDGETNAME, VALUE, XPATH, ACTION, CATEGORY, XPOS, YPOS, VALIDATE]
REQUEST_HEADER_KEYS = [WIDGETNAME, VALUE, XPATH, ACTION, MISC, DYNAMIC, VALIDATE]

# REQUEST_HEADER_PAGEURL_NO = REQUEST_HEADER_KEYS.index(PAGEURL) + 1

REQUEST_HEADER_WIDGETNAME_NO = REQUEST_HEADER_KEYS.index(WIDGETNAME) + 1

REQUEST_HEADER_VALUE_NO = REQUEST_HEADER_KEYS.index(VALUE) + 1

REQUEST_HEADER_XPATH_NO = REQUEST_HEADER_KEYS.index(XPATH) + 1

REQUEST_HEADER_ACTION_NO = REQUEST_HEADER_KEYS.index(ACTION) + 1

# REQUEST_HEADER_CATEGORY_NO = REQUEST_HEADER_KEYS.index(CATEGORY) + 1

# REQUEST_HEADER_XPOS_NO = REQUEST_HEADER_KEYS.index(XPOS) + 1

# REQUEST_HEADER_YPOS_NO = REQUEST_HEADER_KEYS.index(YPOS) + 1

REQUEST_HEADER_VALIDATE_NO = REQUEST_HEADER_KEYS.index(VALIDATE) + 1

REQUEST_HEADER_MISC_NO = REQUEST_HEADER_KEYS.index(MISC) + 1

REQUEST_HEADER_DYNAMIC_NO = REQUEST_HEADER_KEYS.index(DYNAMIC) + 1

# Response
# RESPONSE_HEADER_KEYS = [PAGEURL, WIDGETNAME, VALUE, XPATH, ACTION, CATEGORY, XPOS, YPOS, ACTIVE, EXECUTE]
RESPONSE_HEADER_KEYS = [WIDGETNAME, VALUE, XPATH, ACTION, MISC, DYNAMIC, ACTIVE, EXECUTE]

# RESPONSE_HEADER_PAGEURL_NO = RESPONSE_HEADER_KEYS.index(PAGEURL) + 1

RESPONSE_HEADER_WIDGETNAME_NO = RESPONSE_HEADER_KEYS.index(WIDGETNAME) + 1

RESPONSE_HEADER_VALUE_NO = RESPONSE_HEADER_KEYS.index(VALUE) + 1

RESPONSE_HEADER_XPATH_NO = RESPONSE_HEADER_KEYS.index(XPATH) + 1

RESPONSE_HEADER_ACTION_NO = RESPONSE_HEADER_KEYS.index(ACTION) + 1

# RESPONSE_HEADER_CATEGORY_NO = RESPONSE_HEADER_KEYS.index(CATEGORY) + 1

# RESPONSE_HEADER_XPOS_NO = RESPONSE_HEADER_KEYS.index(XPOS) + 1

# RESPONSE_HEADER_YPOS_NO = RESPONSE_HEADER_KEYS.index(YPOS) + 1

RESPONSE_HEADER_ACTIVE_NO = RESPONSE_HEADER_KEYS.index(ACTIVE) + 1

RESPONSE_HEADER_EXECUTE_NO = RESPONSE_HEADER_KEYS.index(EXECUTE) + 1

RESPONSE_HEADER_MISC_NO = RESPONSE_HEADER_KEYS.index(MISC) + 1

RESPONSE_HEADER_DYNAMIC_NO = RESPONSE_HEADER_KEYS.index(DYNAMIC) + 1

req_headers_length = len(REQUEST_HEADER_KEYS)
resp_headers_length = len(RESPONSE_HEADER_KEYS)
no_of_extra_rows = 7
stepOneRowNumber = 10
extraCols = 2  # Accounts for extra columns A & B (Request Items Count Value)


class GraphNodes:

    def __init__(self, excelDir, graphName, outDir, serverURL, labelNum, maxChar, logoLink):
        self.workbook = load_workbook(excelDir, data_only=True)
        self.totalExcelSheetNumber = len(self.workbook.sheetnames)
        self.graphName = graphName
        self.outDir = outDir
        self.serverURL = serverURL
        self.nonActionItemCondition = labelNum
        self.maxCharLimit = maxChar
        self.companyImage = logoLink
        self.pageUrlRequest = []
        self.widgetNameRequest = []
        self.valueRequest = []
        self.xpathRequest = []
        self.actionRequest = []
        self.categoryRequest = []
        self.xposRequest = []
        self.yposRequest = []
        self.validateRequest = []
        self.requestDictionary = {'PageUrl': self.pageUrlRequest, 'WidgetName': self.widgetNameRequest,
                                  'Value': self.valueRequest, 'XPATH': self.xpathRequest,
                                  'Action': self.actionRequest, 'Category': self.categoryRequest,
                                  'XPOS': self.xposRequest, 'YPOS': self.yposRequest,
                                  'Validate': self.validateRequest}
        self.pageUrlResponse = []
        self.widgetNameResponse = []
        self.valueResponse = []
        self.xpathResponse = []
        self.actionResponse = []
        self.categoryResponse = []
        self.xposResponse = []
        self.yposResponse = []
        self.activeResponse = []
        self.executeResponse = []
        self.responseDictionary = {'PageUrl': self.pageUrlResponse, 'WidgetName': self.widgetNameResponse,
                                   'Value': self.valueResponse, 'XPATH': self.xpathResponse,
                                   'Action': self.actionResponse, 'Category': self.categoryResponse,
                                   'XPOS': self.xposResponse, 'YPOS': self.yposResponse,
                                   'Active': self.activeResponse, 'Execute': self.executeResponse}
        self.mainNodeArray = []
        self.subNodeArray = []
        self.subFinder = 0
        self.previousNodeArray = []
        self.currentLabelArray = []
        self.mainLabelArray = []
        self.mainNodeList = []
        self.widgetValue = "0"
        self.color = "0"
        self.shape = "0"
        self.label = "0"
        self.mainEdgeDict = {}
        self.mainEdgeDictL = {}
        self.subEdgeDict = {}
        self.xpathValue = {"Widget Value": self.widgetValue, "Color": self.color}
        self.mainNodeDict = {}
        self.mainNodeDictL = {}
        self.subNodeDict = {}
        self.edgeArray = []
        self.graphDict = {"Main Nodes": self.mainNodeArray, "Sub Nodes": self.subNodeArray, "Edges": self.edgeArray}
        self.headerSheetName = []
        self.headerPreReq = []
        self.headerStepNum = []
        self.headerPageNum = []
        self.headerPart = []
        self.htmlHeaderPart = []
        self.checkWidgetDuplicateList = []
        self.widgetClear = False
        self.menuItems = []
        self.checkMenuDuplicateList = []
        self.checkMenuDuplicateListL = []
        self.checkIfMenuUsed = False
        self.invisibleLabel = ''
        self.lastLabel = ''
        self.addToMenu = []
        self.finalNode = ''
        self.checkButtonList = []
        self.buttonClear = False
        self.checkNoRequestRepeats = False
        self.checkXpath = []
        self.labelChunk = 0
        self.labelChunkArray = [0]
        self.labelChunkL = 0
        self.labelChunkArrayL = [0]
        self.labelAdd = 0
        self.labelAddL = 0
        self.labelBreak = True
        self.keepEdge = []
        self.keepEdgeL = []

    def clearArray(self):
        self.pageUrlRequest.clear()
        self.widgetNameRequest.clear()
        self.valueRequest.clear()
        self.xpathRequest.clear()
        self.actionRequest.clear()
        self.categoryRequest.clear()
        self.xposRequest.clear()
        self.yposRequest.clear()
        self.validateRequest.clear()
        self.pageUrlResponse.clear()
        self.widgetNameResponse.clear()
        self.valueResponse.clear()
        self.xpathResponse.clear()
        self.actionResponse.clear()
        self.categoryResponse.clear()
        self.xposResponse.clear()
        self.yposResponse.clear()
        self.activeResponse.clear()
        self.executeResponse.clear()
        self.checkWidgetDuplicateList.clear()
        self.checkMenuDuplicateList.clear()
        self.checkMenuDuplicateListL.clear()

    def clearDict(self):
        self.mainNodeDict.clear()
        self.mainEdgeDict.clear()
        self.subNodeDict.clear()
        self.subEdgeDict.clear()

    def sort(self):
        mergedPdf = PdfFileMerger()
        for sheetIndex in range(self.totalExcelSheetNumber):
            if sheetIndex > 0:
                self.clearArray()
            ws = self.workbook[self.workbook.sheetnames[sheetIndex]]
            noOfStepValue = ws.cell(no_of_extra_rows + 1, 2).value
            noOfStep = int(noOfStepValue.split('=')[1])
            numberOfDataset = 1
            label = "A"
            if sheetIndex == 0:
                for header in range(self.totalExcelSheetNumber):
                    ws_header = self.workbook[self.workbook.sheetnames[header]]
                    version = ws_header.cell(5, 2).value
                    software = ws_header.cell(5, 3).value
                    createdOn = ws_header.cell(6, 2).value
                    self.headerSheetName.append(str(ws_header.cell(1, 2).value))
                    self.headerPreReq.append(str(ws_header.cell(7, 2).value))
                    self.headerStepNum.append(str(ws_header.cell(8, 2).value).split('=')[1])
                    self.headerPageNum.append(str(header + 2))
                self.headerSheetName.append(f"{str(self.headerSheetName[-1])} continued")
                self.headerPreReq.append(f"{str(self.headerPreReq[-1])} continued")
                self.headerStepNum.append(f"{str(self.headerStepNum[-1])}")
                dirName = os.path.dirname(__file__)
                # determine if application is a script file or frozen exe
                if getattr(sys, 'frozen', False):
                    application_path = os.path.dirname(sys.executable)
                elif __file__:
                    application_path = os.path.dirname(__file__)
                if self.companyImage != 'Default':
                    fileName = 'testImage_v1'
                    fullPath = f'{application_path}/temp/{fileName}'
                    urllib.request.urlretrieve(self.companyImage, fullPath)
                    img = Image.open(fullPath)
                    rgb_im = img.convert('RGB')
                    rgb_im.save(f'{application_path}/temp/{fileName}.jpg')
                    fileNameFullLogo = f'{application_path}/temp/{fileName}.jpg'
                    fullLogoAddressFull = fileNameFullLogo
                else:
                    fileNameFullLogo = os.path.join(application_path, 'images', 'fullLogo.jpg')
                    fullLogoAddressFull = fileNameFullLogo
                fileNameSmallLogo = os.path.join(application_path, 'images', 'smallLogo.jpg')
                fullLogoAddressSmall = fileNameSmallLogo

            for dataNum in range(numberOfDataset):
                datasetHeaderRowNum = int(no_of_extra_rows) + 1 + ((2 * int(noOfStep) + 1) * int(dataNum))
                for step in range(noOfStep):

                    edgeDict = {}
                    # if sheetIndex > 0:
                    #     step += 1
                    stepHeadersRow = datasetHeaderRowNum + 1 + (2 * int(step))
                    if stepHeadersRow >= ws.max_row:
                        break
                    verificationValue = ws.cell(stepHeadersRow, 2).value
                    # print(f'Verification, "{verificationValue}" at sheet: {sheetIndex + 1}, at step: {step + 1}')
                    if 'CrossVerify' in verificationValue:
                        # print(
                        #     f'Catch Cross Verify, "{verificationValue}" at sheet: {sheetIndex + 1}, at step: {step + 1}\n')
                        continue
                    stepValuesRow = stepHeadersRow + 1
                    requestItemsCount = int(ws.cell(stepValuesRow, 2).value)
                    requestItemsCell = 2
                    # Start at requestItemsCell, check the request items, then the following +1 cell is the response
                    responseItemsCell = requestItemsCell + (requestItemsCount * req_headers_length) + 1
                    responseItemsCount = int(ws.cell(stepValuesRow, responseItemsCell).value)
                    self.mainNodeArray.append(ws.cell(stepOneRowNumber,
                                                      REQUEST_HEADER_VALUE_NO + extraCols).value.replace(">",
                                                                                                         " ").replace(
                        ">>", " ").replace("&", f"and").replace("http", f"\'_http") + " " +
                                              ws.cell(stepOneRowNumber,
                                                      REQUEST_HEADER_ACTION_NO + extraCols).value.replace(">",
                                                                                                          " ").replace(
                                                  ">>", " ").replace("&", f"and").replace("http", f"\'_http"))
                    openUrlXpath = ws.cell(stepOneRowNumber,
                                           REQUEST_HEADER_XPATH_NO + extraCols).value

                    if openUrlXpath is None or " ":
                        openUrlXpath = ws.cell(stepOneRowNumber,
                                               REQUEST_HEADER_VALUE_NO + extraCols).value
                    firstDict = {}
                    firstDict["Widget"] = self.splitString(
                        ws.cell(stepOneRowNumber, REQUEST_HEADER_VALUE_NO + extraCols).value.replace("&",
                                                                                                     "and").replace(">",
                                                                                                                    " ").replace(
                            ">>", " ").replace("&", f"and").replace("http", f"\'_http") + " " + ws.cell(
                            stepOneRowNumber, REQUEST_HEADER_ACTION_NO + extraCols).value.replace("&", "and").replace(
                            ">", " ").replace(
                            ">>", " ").replace("&", f"and").replace("http", f"\'_http"), False)
                    previousLabel = "A0"
                    firstDict["Label"] = previousLabel
                    firstDict["color"] = "black"
                    firstDict["shape"] = "note"
                    self.mainNodeDict[openUrlXpath] = firstDict
                    nodeCounterResponse = 1
                    nodeCounterRequest = 1
                    nodeCounter = 1
                    if step > 0:
                        requestXpath = ws.cell(stepValuesRow, requestItemsCell + (req_headers_length * (
                                requestItemsCount - 1)) + REQUEST_HEADER_XPATH_NO).value
                        for reqPull in range(requestItemsCount):
                            miscVal = js.loads(ws.cell(stepValuesRow, requestItemsCell + (
                                    req_headers_length * reqPull) + REQUEST_HEADER_MISC_NO).value)
                            self.pageUrlRequest.append(miscVal[PAGEURL])
                            self.widgetNameRequest.append(ws.cell(stepValuesRow, requestItemsCell + (
                                    req_headers_length * reqPull) + REQUEST_HEADER_WIDGETNAME_NO).value)
                            self.valueRequest.append(ws.cell(stepValuesRow, requestItemsCell + (
                                    req_headers_length * reqPull) + REQUEST_HEADER_VALUE_NO).value)
                            self.xpathRequest.append(ws.cell(stepValuesRow, requestItemsCell + (
                                    req_headers_length * reqPull) + REQUEST_HEADER_XPATH_NO).value)
                            sheet.spell(str(self.widgetNameRequest[reqPull]).replace("’", "'").replace("‑", "-"),
                                        self.headerSheetName[sheetIndex],
                                        stepValuesRow, requestItemsCell + (
                                                req_headers_length * reqPull) + REQUEST_HEADER_WIDGETNAME_NO,
                                        sheetIndex + 1, self.xpathRequest[reqPull])
                            sheet.spell(str(self.valueRequest[reqPull]).replace("’", "'").replace("‑", "-"),
                                        self.headerSheetName[sheetIndex],
                                        stepValuesRow, requestItemsCell + (
                                                req_headers_length * reqPull) + REQUEST_HEADER_VALUE_NO,
                                        sheetIndex + 1, self.xpathRequest[reqPull])
                            self.actionRequest.append(ws.cell(stepValuesRow, requestItemsCell + (
                                    req_headers_length * reqPull) + REQUEST_HEADER_ACTION_NO).value)
                            self.categoryRequest.append(miscVal[CATEGORY])
                            if reqPull < requestItemsCount - 1:
                                if self.categoryRequest[reqPull] == "MenuLink":
                                    self.shape = "cds"
                                    self.color = "green"
                                elif self.categoryRequest[reqPull] == "FormElement":
                                    self.shape = "note"
                                    self.color = "orange"
                                elif self.categoryRequest[reqPull] == "LabelsText":
                                    self.shape = "invhouse"
                                    self.color = "black"
                                else:
                                    self.shape = "underline"
                                    self.color = "grey11"
                                simpleWidget = self.widgetNameRequest[reqPull].strip().replace('\\', '').replace(">",
                                                                                                                 " ").replace(
                                    ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                simpleValue = self.valueRequest[reqPull].strip().replace('\\', '').replace(">",
                                                                                                           " ").replace(
                                    ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                self.subNodeArray.append(self.widgetNameRequest[reqPull])
                                if self.mainNodeDict.get(self.xpathRequest[reqPull] + self.pageUrlRequest[reqPull]):
                                    tempLabel = self.mainNodeDict.get(
                                        self.xpathRequest[reqPull] + self.pageUrlRequest[reqPull]).get('Label')
                                    keyListEdge = list(self.mainEdgeDict.keys())
                                    for indexEdge in range(len(keyListEdge)):
                                        mainEdge = keyListEdge[indexEdge]
                                        if tempLabel in mainEdge:
                                            self.mainEdgeDict.pop(mainEdge)
                                    self.mainNodeDict.pop(self.xpathRequest[reqPull] + self.pageUrlRequest[reqPull])
                                    firstDict = {}
                                    firstDict["Widget"] = self.splitString(simpleWidget, False)
                                    firstDict["Label"] = label + str(nodeCounter)
                                    firstDict["color"] = self.color
                                    firstDict["shape"] = self.shape
                                    self.subNodeDict[
                                        self.xpathRequest[reqPull] + self.pageUrlRequest[reqPull]] = firstDict
                                    edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                    if reqPull == 0:
                                        self.mainEdgeDict[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                    if reqPull > 0:
                                        self.subEdgeDict[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                    previousLabel = label + str(nodeCounter)
                                    nodeCounter += 1
                        try:
                            xpathIndex = self.xpathRequest.index(requestXpath)
                            self.subNodeArray.append(self.widgetNameRequest[xpathIndex])
                            if self.mainNodeDict.get(self.xpathRequest[reqPull] + self.pageUrlRequest[reqPull]):
                                tempLabel = self.mainNodeDict.get(
                                    self.xpathRequest[reqPull] + self.pageUrlRequest[reqPull]).get('Label')
                                keyListEdge = list(self.mainEdgeDict.keys())
                                for indexEdge in range(len(keyListEdge)):
                                    mainEdge = keyListEdge[indexEdge]
                                    if tempLabel in mainEdge:
                                        self.mainEdgeDict.pop(mainEdge)
                                self.mainNodeDict.pop(self.xpathRequest[reqPull] + self.pageUrlRequest[reqPull])
                                self.subFinder = reqPull
                                if self.categoryRequest[reqPull] == "MenuLink":
                                    self.shape = "cds"
                                    self.color = "green"
                                elif self.categoryRequest[reqPull] == "FormElement":
                                    self.shape = "note"
                                    self.color = "orange"
                                elif self.categoryRequest[reqPull] == "LabelsText":
                                    self.shape = "invhouse"
                                    self.color = "black"
                                else:
                                    self.shape = "underline"
                                    self.color = "grey11"
                                simpleWidget = self.widgetNameRequest[reqPull].strip().replace('\\', '').replace(">",
                                                                                                                 " ").replace(
                                    ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                simpleValue = self.valueRequest[reqPull].strip().replace('\\', '').replace(">",
                                                                                                           " ").replace(
                                    ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                self.subNodeArray.append(self.widgetNameRequest[reqPull])
                                firstDict = {}
                                firstDict["Widget"] = self.splitString(simpleWidget + " " + simpleValue, False)
                                firstDict["Label"] = label + str(nodeCounter)
                                firstDict["color"] = self.color
                                firstDict["shape"] = self.shape
                                self.subNodeDict[self.xpathRequest[reqPull] + self.pageUrlRequest[reqPull]] = firstDict
                                edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                self.subEdgeDict[edgeKey] = [previousLabel, label + str(nodeCounter)]
                        except ValueError:
                            break
                        previousLabel = label + str(nodeCounter)
                        nodeCounter += 1
                    self.checkXpath = self.xpathRequest
                    self.clearArray()
                    for respPull in range(responseItemsCount):
                        cellValueTemp = requestItemsCell + (requestItemsCount * req_headers_length) + (
                                resp_headers_length * respPull) + 1
                        if respPull == 0 and step > 0:
                            newLabel = previousLabel
                            miscVal = js.loads(ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_MISC_NO).value)
                            openUrlXpath = miscVal[PAGEURL]
                            firstDict = {}
                            firstDict["Widget"] = self.splitString(
                                miscVal[PAGEURL].replace("&", "and").replace(">", " ").replace(">>", " ").replace("&",
                                                                                                                  f"and").replace(
                                    "http", f"\'_http"), False)
                            newLabel = label + str(nodeCounter)
                            firstDict["Label"] = newLabel
                            firstDict["color"] = "black"
                            firstDict["shape"] = "note"
                            self.mainNodeDict[openUrlXpath] = firstDict
                            edgeKeyNewPage = previousLabel + "_" + newLabel
                            self.mainEdgeDict[edgeKeyNewPage] = [previousLabel, newLabel]
                            nodeCounter += 1
                            previousLabel = newLabel
                            firstDict = {}
                            break
                        miscVal = js.loads(ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_MISC_NO).value)
                        self.pageUrlResponse.append(miscVal[PAGEURL])
                        self.widgetNameResponse.append(
                            ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_WIDGETNAME_NO).value)
                        self.valueResponse.append(
                            ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_VALUE_NO).value)
                        self.xpathResponse.append(
                            ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_XPATH_NO).value)
                        sheet.spell(str(self.widgetNameResponse[respPull]).replace("’", "'").replace("‑", "-"),
                                    self.headerSheetName[sheetIndex],
                                    stepValuesRow, cellValueTemp + RESPONSE_HEADER_WIDGETNAME_NO,
                                    sheetIndex + 1, self.xpathResponse[respPull])
                        sheet.spell(str(self.valueResponse[respPull]).replace("’", "'").replace("‑", "-"),
                                    self.headerSheetName[sheetIndex],
                                    stepValuesRow, cellValueTemp + RESPONSE_HEADER_VALUE_NO,
                                    sheetIndex + 1, self.xpathResponse[respPull])
                        self.actionResponse.append(
                            ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_ACTION_NO).value)
                        self.activeResponse.append(
                            ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_ACTIVE_NO).value)
                        self.executeResponse.append(
                            ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_EXECUTE_NO).value)
                        self.categoryResponse.append(miscVal[CATEGORY])
                        self.xposResponse.append(miscVal[XPOS])
                        self.yposResponse.append(miscVal[YPOS])
                        if str(self.executeResponse[respPull]).lower() == "true":
                            # Menu Here
                            # if float(self.yposResponse[respPull]) == 0.00 and str(
                            #         self.actionResponse[respPull]).lower() == 'clickbutton' and str(
                            #     self.categoryResponse[respPull]).lower() == 'menulink' and str(
                            #     self.activeResponse[respPull]).lower() == 'true':
                            #     for x in range(len(self.menuItems)):
                            #         if self.valueResponse[respPull]:
                            #             if self.menuItems[x] == self.valueResponse[respPull]:
                            #                 self.checkIfMenuUsed = True
                            #                 # print(f"Found value: {self.valueResponse[respPull]}")
                            #                 break
                            #             else:
                            #                 self.checkIfMenuUsed = False
                            #     self.menuItems.append(f"{self.valueResponse[respPull]}")
                            #     self.shape = "invhouse"
                            #     self.color = "blue"
                            #     simpleWidget = self.widgetNameResponse[respPull].strip().replace(">", " ").replace(
                            #         ">>", " ").replace('"', "").replace(">", " ").replace(
                            #         ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                            #     simpleValue = self.valueResponse[respPull].strip().replace(">", " ").replace(">>",
                            #                                                                                  " ").replace(
                            #         '"', "").replace(">", " ").replace(
                            #         ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                            #     if simpleWidget == simpleValue:
                            #         widgetValue = simpleValue
                            #     elif simpleValue:
                            #         widgetValue = simpleValue + " " + simpleWidget
                            #     else:
                            #         widgetValue = simpleWidget
                            #     for x in range(len(self.checkMenuDuplicateList)):
                            #         if self.checkMenuDuplicateList[x] == widgetValue:
                            #             self.widgetClear = False
                            #             break
                            #         else:
                            #             self.widgetClear = True
                            #     self.checkMenuDuplicateList.append(widgetValue)
                            #     # widgetValue = self.splitString(widgetValue, True)
                            #     widgetValue = self.tableMiddle(widgetValue)
                            #     if self.mainNodeDict.get(
                            #             "sheet" + str(sheetIndex) + "_step" + str(
                            #                 step) + "_labelsMenu") is None and self.widgetClear is True:
                            #         firstDict = {}
                            #         firstDict["Widget"] = widgetValue
                            #         firstDict["Label"] = label + str(nodeCounter)
                            #         firstDict["color"] = self.color
                            #         firstDict["shape"] = self.shape
                            #         self.mainNodeDict[
                            #             "sheet" + str(sheetIndex) + "_step" + str(step) + "_labelsMenu"] = firstDict
                            #         edgeKey = previousLabel + "_" + label + str(nodeCounter)
                            #         self.mainEdgeDict[edgeKey] = [previousLabel, label + str(nodeCounter)]
                            #         nodeCounter += 1
                            #     elif self.widgetClear is True:
                            #         # print(f"{widgetValue}")
                            #         self.mainNodeDict["sheet" + str(sheetIndex) + "_step" + str(step) + "_labelsMenu"][
                            #             "Widget"] = self.mainNodeDict.get(
                            #             "sheet" + str(sheetIndex) + "_step" + str(step) + "_labelsMenu").get(
                            #             "Widget") + " " + widgetValue
                            for x in range(len(self.menuItems)):
                                if self.valueResponse[respPull]:
                                    if self.menuItems[x] == self.valueResponse[respPull]:
                                        self.checkIfMenuUsed = True
                                        # print(f"Found value: {self.valueResponse[respPull]}")
                                        break
                                    else:
                                        self.checkIfMenuUsed = False
                            for x in range(len(self.checkXpath)):
                                # print(f"x: {self.checkXpath[x]}")
                                if str(self.checkXpath[x]) == str(self.xpathResponse[respPull]):
                                    self.checkNoRequestRepeats = False
                                    break
                                else:
                                    self.checkNoRequestRepeats = True
                            self.checkXpath.clear()
                            if self.nonActionItemCondition == 1 and self.checkIfMenuUsed is False:
                                if str(self.activeResponse[respPull]).lower() == "true":
                                    simpleWidget = self.widgetNameResponse[respPull].strip().replace('\\', '').replace(
                                        ">", " ").replace(
                                        ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                    simpleValue = self.valueResponse[respPull].strip().replace('\\', '').replace(">",
                                                                                                                 " ").replace(
                                        ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                    if self.categoryResponse[respPull] == "MenuLink":
                                        self.shape = "cds"
                                        self.color = "green"
                                    elif self.categoryResponse[respPull] == "FormElement":
                                        self.shape = "note"
                                        self.color = "orange"
                                    elif self.categoryResponse[respPull] == "LabelsText":
                                        self.shape = "invhouse"
                                        self.color = "black"
                                    else:
                                        self.shape = "underline"
                                        self.color = "grey11"
                                    if self.mainNodeDict.get(
                                            self.xpathResponse[respPull] + self.pageUrlResponse[
                                                respPull]) is None:
                                        if self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]) is None:
                                            firstDict = {}
                                            if simpleWidget == simpleValue:
                                                self.mainNodeArray.append(self.valueResponse[respPull])
                                                firstDict["Widget"] = self.splitString(self.valueResponse[respPull],
                                                                                       False)
                                            elif simpleValue:
                                                self.mainNodeArray.append(
                                                    self.valueResponse[respPull] + " " + self.widgetNameResponse[
                                                        respPull])
                                                firstDict["Widget"] = self.splitString(
                                                    self.valueResponse[respPull] + " " + self.widgetNameResponse[
                                                        respPull], False)
                                            else:
                                                self.mainNodeArray.append(self.widgetNameResponse[respPull])
                                                firstDict["Widget"] = self.splitString(
                                                    self.widgetNameResponse[respPull], False)
                                            firstDict["Label"] = label + str(nodeCounter)
                                            firstDict["color"] = self.color
                                            firstDict["shape"] = self.shape
                                            self.mainNodeDict[
                                                self.xpathResponse[respPull] + self.pageUrlResponse[
                                                    respPull]] = firstDict
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            self.mainEdgeDict[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                            nodeCounter += 1
                                        else:
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            edgeKey2 = previousLabel + "_" + self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                'Label')
                                            self.subEdgeDict[edgeKey2] = [previousLabel, self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                'Label')]
                                    else:
                                        edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                        if self.mainNodeDict.get(self.xpathResponse[respPull]):
                                            edgeKey2 = previousLabel + "_" + self.mainNodeDict.get(
                                                self.xpathResponse[respPull]).get('Label')
                                            self.mainEdgeDict[edgeKey2] = [previousLabel, self.mainNodeDict.get(
                                                self.xpathResponse[respPull]).get('Label')]
                            elif self.nonActionItemCondition == 4 and self.checkIfMenuUsed is False:
                                if str(self.activeResponse[respPull]).lower() == "true" and str(
                                        self.actionResponse[respPull]).lower() == "clickbutton":
                                    simpleWidget = self.widgetNameResponse[respPull].strip().replace('\\', '').replace(
                                        ">", " ").replace(
                                        ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                    simpleValue = self.valueResponse[respPull].strip().replace('\\', '').replace(">",
                                                                                                                 " ").replace(
                                        ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                    if self.categoryResponse[respPull] == "MenuLink":
                                        self.shape = "cds"
                                        self.color = "green"
                                    elif self.categoryResponse[respPull] == "FormElement":
                                        self.shape = "note"
                                        self.color = "orange"
                                    elif self.categoryResponse[respPull] == "LabelsText":
                                        self.shape = "invhouse"
                                        self.color = "black"
                                    else:
                                        self.shape = "underline"
                                        self.color = "grey11"
                                    if self.mainNodeDict.get(
                                            self.xpathResponse[respPull] + self.pageUrlResponse[
                                                respPull]) is None:
                                        if self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]) is None:
                                            firstDict = {}
                                            if simpleWidget == simpleValue:
                                                self.mainNodeArray.append(self.valueResponse[respPull])
                                                firstDict["Widget"] = self.splitString(self.valueResponse[respPull],
                                                                                       False)
                                            elif simpleValue:
                                                self.mainNodeArray.append(
                                                    self.valueResponse[respPull] + " " + self.widgetNameResponse[
                                                        respPull])
                                                firstDict["Widget"] = self.splitString(
                                                    self.valueResponse[respPull] + " " + self.widgetNameResponse[
                                                        respPull], False)
                                            else:
                                                self.mainNodeArray.append(self.widgetNameResponse[respPull])
                                                firstDict["Widget"] = self.splitString(
                                                    self.widgetNameResponse[respPull], False)
                                            firstDict["Label"] = label + str(nodeCounter)
                                            firstDict["color"] = self.color
                                            firstDict["shape"] = self.shape
                                            self.mainNodeDict[
                                                self.xpathResponse[respPull] + self.pageUrlResponse[
                                                    respPull]] = firstDict
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            self.mainEdgeDict[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                            nodeCounter += 1
                                        else:
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            edgeKey2 = previousLabel + "_" + self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                'Label')
                                            self.subEdgeDict[edgeKey2] = [previousLabel, self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                'Label')]
                                    else:
                                        edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                        if self.mainNodeDict.get(self.xpathResponse[respPull]):
                                            edgeKey2 = previousLabel + "_" + self.mainNodeDict.get(
                                                self.xpathResponse[respPull]).get('Label')
                                            self.mainEdgeDict[edgeKey2] = [previousLabel, self.mainNodeDict.get(
                                                self.xpathResponse[respPull]).get('Label')]
                            elif self.nonActionItemCondition == 2 and self.checkIfMenuUsed is False:
                                # Cond-1
                                if str(self.activeResponse[respPull]).lower() != "true":
                                    self.shape = "box"
                                    self.color = "grey11"
                                    simpleWidget = self.widgetNameResponse[respPull].strip().replace(">", " ").replace(
                                        ">>", " ").replace("&", f"and").replace("Â", "")
                                    simpleValue = self.valueResponse[respPull].strip().replace(">", " ") \
                                        .replace(">>", " ").replace("&", f"and").replace("Â", "")
                                    if simpleWidget == simpleValue:
                                        widgetValue = simpleValue
                                    elif simpleValue:
                                        widgetValue = simpleValue + " " + simpleWidget
                                    else:
                                        widgetValue = simpleWidget
                                    for x in range(len(self.checkWidgetDuplicateList)):
                                        if self.checkWidgetDuplicateList[x] == widgetValue and str(
                                                self.actionResponse[respPull]).lower() != 'settext':
                                            self.widgetClear = False
                                            break
                                        elif str(self.actionResponse[respPull]).lower() != 'settext':
                                            self.widgetClear = True
                                    self.checkWidgetDuplicateList.append(widgetValue)
                                    # widgetValue = self.splitString(widgetValue, True)
                                    widgetValue = self.tableMiddle(widgetValue)
                                    if self.labelAdd == 5:
                                        self.labelChunk += 1
                                        self.labelAdd = 0
                                        self.labelChunkArray.append(self.labelChunk)
                                    if self.mainNodeDict.get(
                                            "sheet" + str(sheetIndex) + "_step" + str(
                                                step) + "_labels" + str(
                                                self.labelChunk)) is None and self.widgetClear is True and str(
                                        self.actionResponse[respPull]).lower() != 'settext':
                                        firstDict = {}
                                        # print(f"{widgetValue}")
                                        # Here 1
                                        firstDict["Widget"] = widgetValue
                                        # print(f"First dict widget: {firstDict['Widget']}")
                                        firstDict["Label"] = label + str(nodeCounter)
                                        firstDict["color"] = self.color
                                        firstDict["shape"] = self.shape
                                        self.mainNodeDict[
                                            "sheet" + str(sheetIndex) + "_step" + str(step) + "_labels" + str(
                                                self.labelChunk)] = firstDict
                                        edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                        self.keepEdge.append(edgeKey)
                                        # print(f"edge: {edgeKey}")
                                        self.mainEdgeDict[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                        # print(f"main inside: {self.mainEdgeDict}")
                                        nodeCounter += 1
                                        self.labelAdd += 1
                                    elif self.widgetClear is True and str(
                                            self.actionResponse[respPull]).lower() != 'settext':
                                        # print(f"{widgetValue}")
                                        self.labelBreak = False
                                        self.mainNodeDict[
                                            "sheet" + str(sheetIndex) + "_step" + str(step) + "_labels" + str(
                                                self.labelChunk)][
                                            "Widget"] = self.mainNodeDict.get(
                                            "sheet" + str(sheetIndex) + "_step" + str(step) + "_labels" + str(
                                                self.labelChunk)).get(
                                            "Widget") + " " + widgetValue
                                        self.labelAdd += 1
                                # elif str(self.actionResponse[
                                #              respPull]).lower() == 'clickbutton' and self.checkNoRequestRepeats is True:
                                #     self.shape = "cds"
                                #     self.color = "green"
                                #     simpleWidget = self.widgetNameResponse[respPull].strip().replace(">", " ").replace(
                                #         ">>", " ").replace("&", f"and")
                                #     simpleValue = self.valueResponse[respPull].strip().replace(">", " ") \
                                #         .replace(">>", " ").replace("&", f"and")
                                #     # print(f"Hit at sheet: {sheetIndex + 1}, step: {step},\n"
                                #     #       f"Widget: {simpleWidget}, Value: {simpleValue}")
                                #     if simpleWidget == simpleValue:
                                #         widgetValue = simpleValue
                                #     elif simpleValue:
                                #         widgetValue = simpleValue + " " + simpleWidget
                                #     else:
                                #         widgetValue = simpleWidget
                                #     for x in range(len(self.checkButtonList)):
                                #         if self.checkButtonList[x] == widgetValue:
                                #             self.buttonClear = False
                                #             break
                                #         else:
                                #             self.buttonClear = True
                                #     self.checkButtonList.append(widgetValue)
                                #     # widgetValue = self.splitString(widgetValue, True)
                                #     widgetValue = self.tableMiddle(widgetValue)
                                #     if self.mainNodeDict.get(
                                #             "sheet" + str(sheetIndex) + "_step" + str(
                                #                 step) + "_labelButton") is None and self.buttonClear is True:
                                #         firstDict = {}
                                #         # print(f"{widgetValue}")
                                #         # Here 1
                                #         firstDict["Widget"] = widgetValue
                                #         # print(f"First dict widget: {firstDict['Widget']}")
                                #         firstDict["Label"] = label + str(nodeCounter)
                                #         firstDict["color"] = self.color
                                #         firstDict["shape"] = self.shape
                                #         self.mainNodeDict[
                                #             "sheet" + str(sheetIndex) + "_step" + str(
                                #                 step) + "_labelButton"] = firstDict
                                #         edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                #         self.mainEdgeDict[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                #         nodeCounter += 1
                                #     elif self.buttonClear is True:
                                #         # print(f"{widgetValue}")
                                #         self.mainNodeDict[
                                #             "sheet" + str(sheetIndex) + "_step" + str(step) + "_labelButton"][
                                #             "Widget"] = self.mainNodeDict.get(
                                #             "sheet" + str(sheetIndex) + "_step" + str(step) + "_labelButton").get(
                                #             "Widget") + " " + widgetValue
                                elif str(self.actionResponse[respPull]).lower() != 'gettext':
                                    simpleWidget = self.widgetNameResponse[respPull].strip().replace('\\', '').replace(
                                        ">", " ").replace(
                                        ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                    simpleValue = self.valueResponse[respPull].strip().replace('\\', '').replace(">",
                                                                                                                 " ").replace(
                                        ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                    if self.categoryResponse[respPull] == "MenuLink":
                                        self.shape = "cds"
                                        self.color = "green"
                                    elif self.categoryResponse[respPull] == "FormElement":
                                        self.shape = "note"
                                        self.color = "orange"
                                    elif self.categoryResponse[respPull] == "LabelsText":
                                        self.shape = "invhouse"
                                        self.color = "black"
                                    else:
                                        self.shape = "underline"
                                        self.color = "grey11"
                                    if self.mainNodeDict.get(
                                            self.xpathResponse[respPull] + self.pageUrlResponse[
                                                respPull]) is None:  # or " ":
                                        if self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]) is None:
                                            firstDict = {}
                                            if simpleWidget == simpleValue:
                                                self.mainNodeArray.append(self.valueResponse[respPull])
                                                firstDict["Widget"] = self.splitString(self.valueResponse[respPull],
                                                                                       False)
                                            elif simpleValue:
                                                self.mainNodeArray.append(
                                                    self.valueResponse[respPull] + " " + self.widgetNameResponse[
                                                        respPull])
                                                firstDict["Widget"] = self.splitString(
                                                    self.valueResponse[respPull] + " " + self.widgetNameResponse[
                                                        respPull], False)
                                            else:
                                                self.mainNodeArray.append(self.widgetNameResponse[respPull])
                                                firstDict["Widget"] = self.splitString(
                                                    self.widgetNameResponse[respPull], False)
                                            firstDict["Label"] = label + str(nodeCounter)
                                            firstDict["color"] = self.color
                                            firstDict["shape"] = self.shape
                                            self.mainNodeDict[
                                                self.xpathResponse[respPull] + self.pageUrlResponse[
                                                    respPull]] = firstDict
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            self.mainEdgeDict[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                            nodeCounter += 1
                                        else:
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            edgeKey2 = previousLabel + "_" + self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                'Label')
                                            self.subEdgeDict[edgeKey2] = [previousLabel, self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                'Label')]
                                    else:
                                        edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                        if self.mainNodeDict.get(self.xpathResponse[respPull]):
                                            edgeKey2 = previousLabel + "_" + self.mainNodeDict.get(
                                                self.xpathResponse[respPull]).get('Label')
                                            self.mainEdgeDict[edgeKey2] = [previousLabel, self.mainNodeDict.get(
                                                self.xpathResponse[respPull]).get('Label')]
                            ######----######
                            elif self.checkIfMenuUsed is False:
                                simpleWidget = self.widgetNameResponse[respPull].strip().replace('\\', '')
                                simpleValue = self.valueResponse[respPull].strip().replace('\\', '')
                                if self.categoryResponse[respPull] == "MenuLink":
                                    self.shape = "cds"
                                    self.color = "green"
                                elif self.categoryResponse[respPull] == "FormElement":
                                    self.shape = "note"
                                    self.color = "orange"
                                elif self.categoryResponse[respPull] == "LabelsText":
                                    self.shape = "invhouse"
                                    self.color = "black"
                                else:
                                    self.shape = "underline"
                                    self.color = "grey11"
                                if self.mainNodeDict.get(
                                        self.xpathResponse[respPull] + self.pageUrlResponse[
                                            respPull]) is None:
                                    if self.subNodeDict.get(
                                            self.xpathResponse[respPull] + self.pageUrlResponse[respPull]) is None:
                                        firstDict = {}
                                        if simpleWidget == simpleValue:
                                            self.mainNodeArray.append(self.valueResponse[respPull])
                                            firstDict["Widget"] = self.splitString(self.valueResponse[respPull], False)
                                        elif simpleValue:
                                            self.mainNodeArray.append(
                                                self.valueResponse[respPull] + " " + self.widgetNameResponse[respPull])
                                            firstDict["Widget"] = self.splitString(
                                                self.valueResponse[respPull] + " " + self.widgetNameResponse[respPull],
                                                False)
                                        else:
                                            self.mainNodeArray.append(self.widgetNameResponse[respPull])
                                            firstDict["Widget"] = self.splitString(self.widgetNameResponse[respPull],
                                                                                   False)
                                        firstDict["Label"] = label + str(nodeCounter)
                                        firstDict["color"] = self.color
                                        firstDict["shape"] = self.shape
                                        self.mainNodeDict[
                                            self.xpathResponse[respPull] + self.pageUrlResponse[respPull]] = firstDict
                                        edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                        self.mainEdgeDict[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                        nodeCounter += 1
                                    else:
                                        edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                        edgeKey2 = previousLabel + "_" + self.subNodeDict.get(
                                            self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get('Label')
                                        self.subEdgeDict[edgeKey2] = [previousLabel, self.subNodeDict.get(
                                            self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get('Label')]
                                else:
                                    edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                    if self.mainNodeDict.get(self.xpathResponse[respPull]):
                                        edgeKey2 = previousLabel + "_" + self.mainNodeDict.get(
                                            self.xpathResponse[respPull]).get('Label')
                                        self.mainEdgeDict[edgeKey2] = [previousLabel, self.mainNodeDict.get(
                                            self.xpathResponse[respPull]).get('Label')]
                    ###################################################################
                    if sheetIndex == (self.totalExcelSheetNumber - 1) and step == (noOfStep - 1):
                        for respPull in range(responseItemsCount):
                            cellValueTemp = requestItemsCell + (requestItemsCount * req_headers_length) + (
                                    resp_headers_length * respPull) + 1
                            if respPull == 0 and step > 0:
                                # print(f"Here")
                                newLabel = previousLabel
                                miscVal = js.loads(
                                    ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_MISC_NO).value)
                                openUrlXpath = miscVal[PAGEURL]
                                firstDict = {}
                                firstDict["Widget"] = self.splitString(miscVal[PAGEURL].replace("&", "and"), False)
                                newLabel = label + str(nodeCounter)
                                self.invisibleLabel = newLabel
                                # print(f'invis: {self.invisibleLabel}')
                                firstDict["Label"] = newLabel
                                firstDict["color"] = "black"
                                firstDict["shape"] = "note"
                                self.mainNodeDictL[openUrlXpath] = firstDict
                                edgeKeyNewPage = previousLabel + "_" + newLabel
                                nodeCounter += 1
                                previousLabel = newLabel
                                firstDict = {}
                            miscVal = js.loads(ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_MISC_NO).value)
                            self.pageUrlResponse.append(miscVal[PAGEURL])
                            self.widgetNameResponse.append(
                                ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_WIDGETNAME_NO).value)
                            self.valueResponse.append(
                                ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_VALUE_NO).value)
                            self.xpathResponse.append(
                                ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_XPATH_NO).value)
                            sheet.spell(str(self.widgetNameResponse[respPull]).replace("’", "'").replace("‑", "-"),
                                        self.headerSheetName[sheetIndex],
                                        stepValuesRow, cellValueTemp + RESPONSE_HEADER_WIDGETNAME_NO,
                                        sheetIndex + 1, self.xpathResponse[respPull])
                            sheet.spell(str(self.valueResponse[respPull]).replace("’", "'").replace("‑", "-"),
                                        self.headerSheetName[sheetIndex],
                                        stepValuesRow, cellValueTemp + RESPONSE_HEADER_VALUE_NO,
                                        sheetIndex + 1, self.xpathResponse[respPull])
                            self.actionResponse.append(
                                ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_ACTION_NO).value)
                            self.activeResponse.append(
                                ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_ACTIVE_NO).value)
                            self.executeResponse.append(
                                ws.cell(stepValuesRow, cellValueTemp + RESPONSE_HEADER_EXECUTE_NO).value)
                            self.categoryResponse.append(miscVal[CATEGORY])
                            self.xposResponse.append(miscVal[XPOS])
                            self.yposResponse.append(miscVal[YPOS])

                            if str(self.executeResponse[respPull]).lower() == "true":
                                # Menu Here
                                # if float(self.yposResponse[respPull]) == 0.00 and str(
                                #         self.actionResponse[respPull]).lower() == 'clickbutton' and str(
                                #     self.categoryResponse[respPull]).lower() == 'menulink' and str(
                                #     self.activeResponse[respPull]).lower() == 'true':
                                #     self.menuItems.append(f"{self.valueResponse[respPull]}")
                                #     self.shape = "invhouse"
                                #     self.color = "blue"
                                #     simpleWidget = self.widgetNameResponse[respPull].strip().replace(">", " ").replace(
                                #         ">>", " ").replace('"', "").replace(">", " ").replace(
                                #         ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                #     simpleValue = self.valueResponse[respPull].strip().replace(">", " ").replace(">>",
                                #                                                                                  " ").replace(
                                #         '"', "").replace(">", " ").replace(
                                #         ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                #     if simpleWidget == simpleValue:
                                #         widgetValue = simpleValue
                                #     elif simpleValue:
                                #         widgetValue = simpleValue + " " + simpleWidget
                                #     else:
                                #         widgetValue = simpleWidget
                                #     for x in range(len(self.checkMenuDuplicateList)):
                                #         if self.checkMenuDuplicateList[x] == widgetValue:
                                #             self.widgetClear = False
                                #             break
                                #         else:
                                #             self.widgetClear = True
                                #     self.checkMenuDuplicateList.append(widgetValue)
                                #     # widgetValue = self.splitString(widgetValue, True)
                                #     widgetValue = self.tableMiddle(widgetValue)
                                #     if self.mainNodeDictL.get(
                                #             "sheet" + str(sheetIndex) + "_step" + str(
                                #                 step) + "_labelsMenu") is None and self.widgetClear is True:
                                #         firstDict = {}
                                #         firstDict["Widget"] = widgetValue
                                #         firstDict["Label"] = label + str(nodeCounter)
                                #         firstDict["color"] = self.color
                                #         firstDict["shape"] = self.shape
                                #         self.mainNodeDictL[
                                #             "sheet" + str(sheetIndex) + "_step" + str(step) + "_labelsMenu"] = firstDict
                                #         edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                #         self.mainEdgeDictL[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                #         nodeCounter += 1
                                #     elif self.widgetClear is True:
                                #         # print(f"{widgetValue}")
                                #         self.mainNodeDictL[
                                #             "sheet" + str(sheetIndex) + "_step" + str(step) + "_labelsMenu"][
                                #             "Widget"] = self.mainNodeDictL.get(
                                #             "sheet" + str(sheetIndex) + "_step" + str(step) + "_labelsMenu").get(
                                #             "Widget") + " " + widgetValue
                                # print(f"Menu items for sheet: {sheetIndex + 1}\n{self.menuItems}")
                                for x in range(len(self.menuItems)):
                                    if self.valueResponse[respPull]:
                                        if self.menuItems[x] == self.valueResponse[respPull]:
                                            self.checkIfMenuUsed = True
                                            break
                                        else:
                                            self.checkIfMenuUsed = False
                                if self.nonActionItemCondition == 1 and self.checkIfMenuUsed is False:
                                    if str(self.activeResponse[respPull]).lower() == "true":
                                        simpleWidget = self.widgetNameResponse[respPull].strip().replace('\\',
                                                                                                         '').replace(
                                            ">", " ").replace(
                                            ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                        simpleValue = self.valueResponse[respPull].strip().replace('\\', '').replace(
                                            ">", " ").replace(
                                            ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                        if self.categoryResponse[respPull] == "MenuLink":
                                            self.shape = "cds"
                                            self.color = "green"
                                        elif self.categoryResponse[respPull] == "FormElement":
                                            self.shape = "note"
                                            self.color = "orange"
                                        elif self.categoryResponse[respPull] == "LabelsText":
                                            self.shape = "invhouse"
                                            self.color = "black"
                                        else:
                                            self.shape = "underline"
                                            self.color = "grey11"
                                        if self.mainNodeDictL.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[
                                                    respPull]) is None:  # or " ":
                                            if self.subNodeDict.get(
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[
                                                        respPull]) is None:
                                                firstDict = {}
                                                if simpleWidget == simpleValue:
                                                    self.mainNodeArray.append(self.valueResponse[respPull])
                                                    firstDict["Widget"] = self.splitString(self.valueResponse[respPull],
                                                                                           False)
                                                elif simpleValue:
                                                    self.mainNodeArray.append(
                                                        self.valueResponse[respPull] + " " + self.widgetNameResponse[
                                                            respPull])
                                                    firstDict["Widget"] = self.splitString(
                                                        self.valueResponse[respPull] + " " + self.widgetNameResponse[
                                                            respPull], False)
                                                else:
                                                    self.mainNodeArray.append(self.widgetNameResponse[respPull])
                                                    firstDict["Widget"] = self.splitString(
                                                        self.widgetNameResponse[respPull], False)
                                                firstDict["Label"] = label + str(nodeCounter)
                                                firstDict["color"] = self.color
                                                firstDict["shape"] = self.shape
                                                self.mainNodeDictL[
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[
                                                        respPull]] = firstDict
                                                edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                                self.mainEdgeDictL[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                                nodeCounter += 1
                                            else:
                                                edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                                edgeKey2 = previousLabel + "_" + self.subNodeDict.get(
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                    'Label')
                                                self.subEdgeDict[edgeKey2] = [previousLabel, self.subNodeDict.get(
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                    'Label')]
                                        else:
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            if self.mainNodeDictL.get(self.xpathResponse[respPull]):
                                                edgeKey2 = previousLabel + "_" + self.mainNodeDict.get(
                                                    self.xpathResponse[respPull]).get('Label')
                                                self.mainEdgeDictL[edgeKey2] = [previousLabel, self.mainNodeDict.get(
                                                    self.xpathResponse[respPull]).get('Label')]
                                elif self.nonActionItemCondition == 4 and self.checkIfMenuUsed is False:
                                    if str(self.activeResponse[respPull]).lower() == "true" and str(
                                            self.actionResponse[respPull]).lower() == "clickbutton":
                                        simpleWidget = self.widgetNameResponse[respPull].strip().replace('\\',
                                                                                                         '').replace(
                                            ">", " ").replace(
                                            ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                        simpleValue = self.valueResponse[respPull].strip().replace('\\', '').replace(
                                            ">", " ").replace(
                                            ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                        if self.categoryResponse[respPull] == "MenuLink":
                                            self.shape = "cds"
                                            self.color = "green"
                                        elif self.categoryResponse[respPull] == "FormElement":
                                            self.shape = "note"
                                            self.color = "orange"
                                        elif self.categoryResponse[respPull] == "LabelsText":
                                            self.shape = "invhouse"
                                            self.color = "black"
                                        else:
                                            self.shape = "underline"
                                            self.color = "grey11"
                                        if self.mainNodeDictL.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[
                                                    respPull]) is None:  # or " ":
                                            if self.subNodeDict.get(
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[
                                                        respPull]) is None:
                                                firstDict = {}
                                                if simpleWidget == simpleValue:
                                                    self.mainNodeArray.append(self.valueResponse[respPull])
                                                    firstDict["Widget"] = self.splitString(self.valueResponse[respPull],
                                                                                           False)
                                                elif simpleValue:
                                                    self.mainNodeArray.append(
                                                        self.valueResponse[respPull] + " " + self.widgetNameResponse[
                                                            respPull])
                                                    firstDict["Widget"] = self.splitString(
                                                        self.valueResponse[respPull] + " " + self.widgetNameResponse[
                                                            respPull], False)
                                                else:
                                                    self.mainNodeArray.append(self.widgetNameResponse[respPull])
                                                    firstDict["Widget"] = self.splitString(
                                                        self.widgetNameResponse[respPull], False)
                                                firstDict["Label"] = label + str(nodeCounter)
                                                firstDict["color"] = self.color
                                                firstDict["shape"] = self.shape
                                                self.mainNodeDictL[
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[
                                                        respPull]] = firstDict
                                                edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                                self.mainEdgeDictL[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                                nodeCounter += 1
                                            else:
                                                edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                                edgeKey2 = previousLabel + "_" + self.subNodeDict.get(
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                    'Label')
                                                self.subEdgeDict[edgeKey2] = [previousLabel, self.subNodeDict.get(
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                    'Label')]
                                        else:
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            if self.mainNodeDictL.get(self.xpathResponse[respPull]):
                                                edgeKey2 = previousLabel + "_" + self.mainNodeDict.get(
                                                    self.xpathResponse[respPull]).get('Label')
                                                self.mainEdgeDictL[edgeKey2] = [previousLabel, self.mainNodeDict.get(
                                                    self.xpathResponse[respPull]).get('Label')]
                                elif self.nonActionItemCondition == 2 and self.checkIfMenuUsed is False:
                                    # Cond-2
                                    if str(self.activeResponse[respPull]).lower() != "true":
                                        self.shape = "box"
                                        self.color = "grey11"
                                        simpleWidget = self.widgetNameResponse[respPull].strip().replace(">",
                                                                                                         " ").replace(
                                            ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                        simpleValue = self.valueResponse[respPull].strip().replace(">", " ").replace(
                                            ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                        if simpleWidget == simpleValue:
                                            widgetValue = simpleValue
                                        elif simpleValue:
                                            widgetValue = simpleValue + " " + simpleWidget
                                        else:
                                            widgetValue = simpleWidget
                                        for x in range(len(self.checkWidgetDuplicateList)):
                                            if self.checkWidgetDuplicateList[x] == widgetValue and str(
                                                    self.actionResponse[respPull]).lower() != 'settext':
                                                self.widgetClear = False
                                                break
                                            elif str(self.actionResponse[respPull]).lower() != 'settext':
                                                self.widgetClear = True
                                        self.checkWidgetDuplicateList.append(widgetValue)
                                        # widgetValue = self.splitString(widgetValue, True)
                                        widgetValue = self.tableMiddle(widgetValue)
                                        if self.labelAddL == 5:
                                            self.labelChunkL += 1
                                            self.labelAddL = 0
                                            self.labelChunkArrayL.append(self.labelChunkL)
                                        if self.mainNodeDictL.get(
                                                "sheet" + str(sheetIndex) + "_step" + str(
                                                    step) + "_labels" + str(
                                                    self.labelChunkL)) is None and self.widgetClear is True and str(
                                            self.actionResponse[respPull]).lower() != 'settext':
                                            firstDict = {}
                                            # print(f"{widgetValue}")
                                            # Here 2
                                            firstDict["Widget"] = widgetValue
                                            firstDict["Label"] = label + str(nodeCounter)
                                            firstDict["color"] = self.color
                                            firstDict["shape"] = self.shape
                                            self.mainNodeDictL[
                                                "sheet" + str(sheetIndex) + "_step" + str(step) + "_labels" + str(
                                                    self.labelChunkL)] = firstDict
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            self.keepEdgeL = edgeKey
                                            self.mainEdgeDictL[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                            nodeCounter += 1
                                            self.labelAddL += 1
                                        elif self.widgetClear is True and str(
                                                self.actionResponse[respPull]).lower() != 'settext':
                                            # print(f"{widgetValue}")
                                            self.mainNodeDictL[
                                                "sheet" + str(sheetIndex) + "_step" + str(step) + "_labels" + str(
                                                    self.labelChunkL)][
                                                "Widget"] = self.mainNodeDictL.get(
                                                "sheet" + str(sheetIndex) + "_step" + str(step) + "_labels" + str(
                                                    self.labelChunkL)).get(
                                                "Widget") + " " + widgetValue
                                            self.labelAddL += 1

                                    else:
                                        simpleWidget = self.widgetNameResponse[respPull].strip().replace('\\',
                                                                                                         '').replace(
                                            ">", " ").replace(
                                            ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                        simpleValue = self.valueResponse[respPull].strip().replace('\\', '').replace(
                                            ">", " ").replace(
                                            ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                        if self.categoryResponse[respPull] == "MenuLink":
                                            self.shape = "cds"
                                            self.color = "green"
                                        elif self.categoryResponse[respPull] == "FormElement":
                                            self.shape = "note"
                                            self.color = "orange"
                                        elif self.categoryResponse[respPull] == "LabelsText":
                                            self.shape = "invhouse"
                                            self.color = "black"
                                        else:
                                            self.shape = "underline"
                                            self.color = "grey11"
                                        if self.mainNodeDictL.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[
                                                    respPull]) is None:  # or " ":
                                            if self.subNodeDict.get(
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[
                                                        respPull]) is None:
                                                firstDict = {}
                                                if simpleWidget == simpleValue:
                                                    # print(f"Node: {self.valueResponse[respPull]}")
                                                    self.mainNodeArray.append(self.valueResponse[respPull])
                                                    firstDict["Widget"] = self.splitString(self.valueResponse[respPull],
                                                                                           False)
                                                elif simpleValue:
                                                    self.mainNodeArray.append(
                                                        self.valueResponse[respPull] + " " + self.widgetNameResponse[
                                                            respPull])
                                                    firstDict["Widget"] = self.splitString(
                                                        self.valueResponse[respPull] + " " + self.widgetNameResponse[
                                                            respPull], False)
                                                else:
                                                    self.mainNodeArray.append(self.widgetNameResponse[respPull])
                                                    firstDict["Widget"] = self.splitString(
                                                        self.widgetNameResponse[respPull], False)
                                                firstDict["Label"] = label + str(nodeCounter)
                                                firstDict["color"] = self.color
                                                firstDict["shape"] = self.shape
                                                self.mainNodeDictL[
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[
                                                        respPull]] = firstDict
                                                edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                                self.mainEdgeDictL[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                                nodeCounter += 1
                                            else:
                                                edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                                edgeKey2 = previousLabel + "_" + self.subNodeDict.get(
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                    'Label')
                                                self.subEdgeDict[edgeKey2] = [previousLabel, self.subNodeDict.get(
                                                    self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                    'Label')]
                                        else:
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            if self.mainNodeDictL.get(self.xpathResponse[respPull]):
                                                edgeKey2 = previousLabel + "_" + self.mainNodeDictL.get(
                                                    self.xpathResponse[respPull]).get('Label')
                                                self.mainEdgeDictL[edgeKey2] = [previousLabel, self.mainNodeDictL.get(
                                                    self.xpathResponse[respPull]).get('Label')]
                                elif self.checkIfMenuUsed is False:
                                    simpleWidget = self.widgetNameResponse[respPull].strip().replace('\\', '').replace(
                                        ">", " ").replace(
                                        ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                    simpleValue = self.valueResponse[respPull].strip().replace('\\', '').replace(">",
                                                                                                                 " ").replace(
                                        ">>", " ").replace("&", f"and").replace("http", f"\'_http")
                                    if self.categoryResponse[respPull] == "MenuLink":
                                        self.shape = "cds"
                                        self.color = "green"
                                    elif self.categoryResponse[respPull] == "FormElement":
                                        self.shape = "note"
                                        self.color = "orange"
                                    elif self.categoryResponse[respPull] == "LabelsText":
                                        self.shape = "invhouse"
                                        self.color = "black"
                                    else:
                                        self.shape = "underline"
                                        self.color = "grey11"
                                    if self.mainNodeDictL.get(
                                            self.xpathResponse[respPull] + self.pageUrlResponse[
                                                respPull]) is None:  # or " ":
                                        if self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]) is None:
                                            firstDict = {}
                                            if simpleWidget == simpleValue:
                                                # print(f"Node: {self.valueResponse[respPull]}")
                                                self.mainNodeArray.append(self.valueResponse[respPull])
                                                firstDict["Widget"] = self.splitString(self.valueResponse[respPull],
                                                                                       False)
                                            elif simpleValue:
                                                self.mainNodeArray.append(
                                                    self.valueResponse[respPull] + " " + self.widgetNameResponse[
                                                        respPull])
                                                firstDict["Widget"] = self.splitString(
                                                    self.valueResponse[respPull] + " " + self.widgetNameResponse[
                                                        respPull], False)
                                            else:
                                                self.mainNodeArray.append(self.widgetNameResponse[respPull])
                                                firstDict["Widget"] = self.splitString(
                                                    self.widgetNameResponse[respPull], False)
                                            firstDict["Label"] = label + str(nodeCounter)
                                            firstDict["color"] = self.color
                                            firstDict["shape"] = self.shape
                                            self.mainNodeDictL[
                                                self.xpathResponse[respPull] + self.pageUrlResponse[
                                                    respPull]] = firstDict
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            self.mainEdgeDictL[edgeKey] = [previousLabel, label + str(nodeCounter)]
                                            nodeCounter += 1
                                        else:
                                            edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                            edgeKey2 = previousLabel + "_" + self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                'Label')
                                            self.subEdgeDict[edgeKey2] = [previousLabel, self.subNodeDict.get(
                                                self.xpathResponse[respPull] + self.pageUrlResponse[respPull]).get(
                                                'Label')]
                                    else:
                                        edgeKey = previousLabel + "_" + label + str(nodeCounter)
                                        if self.mainNodeDictL.get(self.xpathResponse[respPull]):
                                            edgeKey2 = previousLabel + "_" + self.mainNodeDict.get(
                                                self.xpathResponse[respPull]).get('Label')
                                            self.mainEdgeDictL[edgeKey2] = [previousLabel, self.mainNodeDict.get(
                                                self.xpathResponse[respPull]).get('Label')]
                                    self.lastLabel = edgeKey2
                    ###################################################################
                    # Menu items
                    # saveYpos = None
                    # saveLastPos = 0
                    # tempString = ''
                    # for x in range(len(self.yposResponse)):
                    #     if saveYpos == float(self.yposResponse[x]):
                    #         if str(self.activeResponse[x]).lower() == 'true' and str(
                    #                 self.executeResponse[x]).lower() == 'true' and str(
                    #             self.valueResponse[x]) != '':
                    #             # print(f"Menu Item: {self.valueResponse[x]}\n")
                    #             tempString = self.tableMiddle(str(self.valueResponse[x]))
                    #             self.addToMenu.append(tempString)
                    #             saveLastPos = x
                    #     elif float(self.yposResponse[x]) == float(self.yposResponse[x - 1]):
                    #         if str(self.activeResponse[x - 1]).lower() == 'true' and str(
                    #                 self.executeResponse[x - 1]).lower() == 'true' and str(
                    #             self.valueResponse[x - 1]) != '':
                    #             tempString = self.tableMiddle(str(self.valueResponse[x - 1]))
                    #             self.addToMenu.append(tempString)
                    #             saveYpos = float(self.yposResponse[x - 1])
                    #             saveLastPos = x
                    # if str(self.activeResponse[saveLastPos]).lower() == 'true' and str(
                    #         self.executeResponse[saveLastPos]).lower() == 'true' and str(
                    #     self.valueResponse[saveLastPos]) != '':
                    #     tempString = self.tableMiddle(str(self.valueResponse[saveLastPos]))
                    #     self.addToMenu.append(tempString)
                    # addString = ''
                    # if self.addToMenu:
                    #     for x in range(len(self.addToMenu)):
                    #         addString += self.addToMenu[x]
                    #     self.finalNode = self.addTableStartEnd(self.tableMiddle(' Menu Items:') + addString)
                    # print(f"Final Node: {self.finalNode}")
                    # firstDict = {}
                    # firstDict["Label"] = 'Z0'
                    # firstDict["Color"] = 'blue'
                    # firstDict["Shape"] = 'invhouse'
                    # firstDict["Widget"] = self.finalNode
                    # self.mainNodeDict["sheet" + str(sheetIndex) + "_step" + str(step) + "_labelsMainMenu"] = firstDict
                    # self.mainEdgeDict['A0_Z0'] = ['A0', 'Z0']
                    ###################################################################

                    ###################################################################
                    for x in range(len(self.labelChunkArray)):
                        if self.mainNodeDict.get(
                                "sheet" + str(sheetIndex) + "_step" + str(step) + "_labels" + str(x)) is None:
                            pass
                        else:
                            # print(self.mainNodeDict.get(
                            #     "sheet" + str(sheetIndex) + "_step" + str(step) + "_labels").get("Widget"))
                            tempWidget = self.mainNodeDict.get(
                                "sheet" + str(sheetIndex) + "_step" + str(step) + "_labels" + str(x)).get("Widget")
                            # print(f"Temp widget: {tempWidget}")
                            self.mainNodeDict["sheet" + str(sheetIndex) + "_step" + str(step) + "_labels" + str(x)][
                                "Widget"] = self.addTableStartEnd(self.tableMiddle(' Labels:') + tempWidget)
                    ###################################################################
                    if self.mainNodeDict.get("sheet" + str(sheetIndex) + "_step" + str(step) + "_labelsMenu") is None:
                        pass
                    else:
                        tempWidget = self.mainNodeDict.get(
                            "sheet" + str(sheetIndex) + "_step" + str(step) + "_labelsMenu").get("Widget")
                        self.mainNodeDict["sheet" + str(sheetIndex) + "_step" + str(step) + "_labelsMenu"][
                            "Widget"] = self.addTableStartEnd(self.tableMiddle(' Menu Items:') + tempWidget)
                    ###################################################################
                    if self.mainNodeDict.get(
                            "sheet" + str(sheetIndex) + "_step" + str(step) + "_labelButton") is None:
                        pass
                    else:
                        tempWidget = self.mainNodeDict.get(
                            "sheet" + str(sheetIndex) + "_step" + str(step) + "_labelButton").get("Widget")
                        self.mainNodeDict["sheet" + str(sheetIndex) + "_step" + str(step) + "_labelButton"][
                            "Widget"] = self.addTableStartEnd(self.tableMiddle(' Button Items:') + tempWidget)
                    ###################################################################
                    for y in range(len(self.labelChunkArrayL)):
                        if self.mainNodeDictL.get(
                                "sheet" + str(sheetIndex) + "_step" + str(step) + "_labels" + str(y)) is None:
                            pass
                        else:
                            tempWidget = self.mainNodeDictL.get(
                                "sheet" + str(sheetIndex) + "_step" + str(step) + "_labels" + str(y)).get("Widget")
                            # print(f"Temp widget: {tempWidget}")
                            self.mainNodeDictL[
                                "sheet" + str(sheetIndex) + "_step" + str(step) + "_labels" + str(y)][
                                "Widget"] = self.addTableStartEnd(self.tableMiddle(' Labels:') + tempWidget)
                    ###################################################################
                    if self.mainNodeDictL.get("sheet" + str(sheetIndex) + "_step" + str(step) + "_labelsMenu") is None:
                        pass
                    else:
                        tempWidget = self.mainNodeDictL.get(
                            "sheet" + str(sheetIndex) + "_step" + str(step) + "_labelsMenu").get("Widget")
                        # print(f"Temp widget: {tempWidget}")
                        self.mainNodeDictL["sheet" + str(sheetIndex) + "_step" + str(step) + "_labelsMenu"][
                            "Widget"] = self.addTableStartEnd(self.tableMiddle(' Menu Items:') + tempWidget)
                    ###################################################################
                    previousLabel = label + str(nodeCounter - 1)
                    label = chr(ord(label) + 1)
                    # print(f"Menu items for sheet: {sheetIndex + 1}\n{self.menuItems}")
                    self.menuItems.clear()
                    # print(f'main after clear: {self.mainEdgeDict}')
            # print(f'main before buttons: {self.mainEdgeDict}')
            indexList = []
            keyTest = []
            tempButtons = []
            # print(f"Main edge1: {self.mainEdgeDict}\n////////////\n")
            keyList = list(self.mainNodeDict.keys())
            for index in range(len(keyList)):
                mainNode = keyList[index]
                if self.mainNodeDict.get(mainNode).get('shape') == 'cds':
                    # print(f"\nCDS match at index: {index}")
                    indexList.append(index)
                    keyTest.append(mainNode)
            # print(f"\n//////////\nindexlist: {indexList}\n")
            labelList = []
            for x in range(len(indexList)):
                testNode = keyTest[x]
                # print(f"Shape at index {indexList[x]}: {self.mainNodeDict.get(testNode).get('shape')}\n"
                #       f"Label: {self.mainNodeDict.get(testNode).get('Label')}\n"
                #       f"Key-Value: {self.mainNodeDict.get(testNode)}\n///////\n")
                labelList.append(self.mainNodeDict.get(testNode).get('Label'))
                # print(f"testNode: {testNode}")
                # print(f"Key-val: {self.mainNodeDict.get(testNode).get('Widget')}")
                tempButtons.append(self.mainNodeDict.get(testNode).get('Widget'))
                self.mainNodeDict.pop(testNode)
            # print(f"main node after {self.mainNodeDict}")
            # print(f"Tembuttons: {tempButtons}")
            tempButtons.sort()
            finalWidget = ''
            for x in tempButtons:
                finalWidget += x.replace("<<table border='0' cellborder='0' CELLSPACING='-5'>", "").replace(
                    "</table>>", "")
            # print(f"Final widget: {finalWidget}")
            # print(f"Main edge: {self.mainEdgeDict}\n"
            #       f"Label list: {labelList}")
            keyTemp = []
            for key in self.mainEdgeDict.items():
                for label in labelList:
                    if label in key[0].split('_')[1]:
                        keyTemp.append(key[0])
            # print(f"temp: {keyTemp}")
            # print(f"key: {keyTemp}")
            # print(f"keep: {self.keepEdge}")
            # print(f"edge1: {self.mainEdgeDict}")
            keyTempAfterKeepEdge = [x for x in keyTemp if x not in self.keepEdge]
            for x in keyTempAfterKeepEdge:
                try:
                    self.mainEdgeDict.pop(x)
                except KeyError:
                    pass
            # print(f"edge2: {self.mainEdgeDict}")
            # print(f"main edge after: {self.mainEdgeDict}")
            firstDict = {}
            firstDict["Widget"] = finalWidget
            firstDict["Label"] = 'Z9'
            firstDict["color"] = 'green'
            firstDict["shape"] = 'cds'
            self.mainNodeDict["sheet" + str(sheetIndex) + "_step" + str(step) + "_labelFullButtons"] = firstDict
            edgeKey = "A0_Z9"
            self.mainEdgeDict[edgeKey] = ["A0", "Z9"]
            tempWidget = self.mainNodeDict.get(
                "sheet" + str(sheetIndex) + "_step" + str(step) + "_labelFullButtons").get("Widget")
            self.mainNodeDict["sheet" + str(sheetIndex) + "_step" + str(step) + "_labelFullButtons"][
                "Widget"] = self.addTableStartEnd(self.tableMiddle(' Button Items:') + tempWidget)
            # print(f"Main edge2: {self.mainEdgeDict}\n////////////\n")
            ###################################################################################
            indexListL = []
            keyTestL = []
            tempButtonsL = []
            # print(f"Main edge1: {self.mainEdgeDict}\n////////////\n")
            keyListL = list(self.mainNodeDictL.keys())
            for index in range(len(keyListL)):
                mainNodeL = keyListL[index]
                if self.mainNodeDictL.get(mainNodeL).get('shape') == 'cds':
                    # print(f"\nCDS match at index: {index}")
                    indexListL.append(index)
                    keyTestL.append(mainNodeL)
            # print(f"\n//////////\nindexlist: {indexList}\n")
            labelListL = []
            for x in range(len(indexListL)):
                testNodeL = keyTestL[x]
                # print(f"Shape at index {indexList[x]}: {self.mainNodeDict.get(testNode).get('shape')}\n"
                #       f"Label: {self.mainNodeDict.get(testNode).get('Label')}\n"
                #       f"Key-Value: {self.mainNodeDict.get(testNode)}\n///////\n")
                labelListL.append(self.mainNodeDictL.get(testNodeL).get('Label'))
                # print(f"testNode: {testNode}")
                # print(f"Key-val: {self.mainNodeDict.get(testNode).get('Widget')}")
                tempButtonsL.append(self.mainNodeDictL.get(testNodeL).get('Widget'))
                self.mainNodeDictL.pop(testNodeL)
            # print(f"main node after {self.mainNodeDictL}")
            # print(f"Tembuttons: {tempButtons}")
            tempButtonsL.sort()
            finalWidgetL = ''
            for x in tempButtonsL:
                finalWidgetL += x.replace("<<table border='0' cellborder='0' CELLSPACING='-5'>", "").replace(
                    "</table>>", "")
            # print(f"Final widget: {finalWidget}")
            # print(f"Main edge: {self.mainEdgeDict}\n"
            #       f"Label list: {labelList}")
            keyTempL = []
            for key in self.mainEdgeDictL.items():
                for label in labelListL:
                    if label in key[0].split('_')[1]:
                        keyTempL.append(key[0])
            # print(f"temp: {keyTemp}")
            # print(f"key: {keyTemp}")
            # print(f"keep: {self.keepEdge}")
            # print(f"edge1: {self.mainEdgeDict}")
            keyTempAfterKeepEdgeL = [x for x in keyTempL if x not in self.keepEdgeL]
            for x in keyTempAfterKeepEdgeL:
                try:
                    self.mainEdgeDictL.pop(x)
                except KeyError:
                    pass
            # print(f"edge2: {self.mainEdgeDict}")
            # print(f"main edge after: {self.mainEdgeDict}")
            firstDict = {}
            firstDict["Widget"] = finalWidgetL
            firstDict["Label"] = 'X9'
            firstDict["color"] = 'green'
            firstDict["shape"] = 'cds'
            self.mainNodeDictL["sheet" + str(sheetIndex) + "_step" + str(step) + "_labelFullButtons"] = firstDict
            edgeKeyL = self.invisibleLabel + "_X9"
            self.mainEdgeDictL[edgeKeyL] = [self.invisibleLabel, "X9"]
            tempWidgetL = self.mainNodeDictL.get(
                "sheet" + str(sheetIndex) + "_step" + str(step) + "_labelFullButtons").get("Widget")
            self.mainNodeDictL["sheet" + str(sheetIndex) + "_step" + str(step) + "_labelFullButtons"][
                "Widget"] = self.addTableStartEnd(self.tableMiddle(' Button Items:') + tempWidgetL)
            ###################################################################################
            self.labelChunk = 0
            self.labelChunkArray = [0]
            self.labelChunkL = 0
            self.labelChunkArrayL = [0]
            self.labelAdd = 0
            self.labelAddL = 0
            self.keepEdge.clear()
            self.graph(mergedPdf, sheetIndex)
            # if sheetIndex == (self.totalExcelSheetNumber - 1):
            #     self.graphL(mergedPdf, sheetIndex)
            # sheet.createPDF()
            self.clearDict()
            # break
            sheet.createPDF()
        mergedPdf.write(f"{self.outDir}{self.graphName}.pdf")
        s = time.time()
        add_header_and_footer(f"{self.outDir}{self.graphName}.pdf", self.headerSheetName, self.headerPreReq,
                              self.headerStepNum, fullLogoAddressFull, fullLogoAddressSmall, version, software,
                              createdOn, "spellcheck.pdf")
        e = time.time()
        print("add_header_and_footer: " + str(e - s))
        os.remove(f"{self.outDir}{self.graphName}")
        if self.companyImage != 'Default':
            os.remove(fullPath)
            os.remove(f'{application_path}/temp/{fileName}.jpg')
        print(f"{self.serverURL}\n{self.graphName}.pdf")

    def splitString(self, widgetString, isLabel):
        stringList = []
        maxChar = 60
        widgetString = widgetString.replace("&", "and").replace(">", "").replace("’", "'")
        for x in range(len(textwrap.wrap(widgetString, maxChar, break_long_words=True))):
            stringList.append(textwrap.wrap(widgetString, maxChar, break_long_words=True)[x])
        if len(stringList) == 0:
            # print(f"Empty here 1")
            stringList = "!"
        tableStart = f"<<table border='0' cellborder='0' CELLSPACING='-5'>"
        tableEnd = f"</table>>"
        tableMiddle = ""
        for x in range(len(stringList)):
            if x == 0 and isLabel is True:
                tableMiddle += f"<tr><td ALIGN='LEFT'><FONT POINT-SIZE='11' COLOR='red'><B>{stringList[x][0].capitalize()}</B></FONT>" \
                               f"<FONT POINT-SIZE='10'>{stringList[x][1:]}</FONT></td></tr>"
            else:
                tableMiddle += f"<tr><td ALIGN='LEFT'><FONT POINT-SIZE='10'>{stringList[x]}</FONT></td></tr>"
        fullHTML = tableStart + tableMiddle + tableEnd
        return fullHTML

    def addTableStartEnd(self, tableMiddle):
        tableStart = f"<<table border='0' cellborder='0' CELLSPACING='-5' CELLPADDING='3'>"
        tableEnd = f"</table>>"
        fullHTML = tableStart + tableMiddle + tableEnd
        return fullHTML

    def tableMiddle(self, widgetString):
        stringList = []
        maxChar = 60
        tableMiddle = ''
        widgetString = widgetString.replace("&", "and").replace(">", "").replace("’", "'")
        for x in range(len(textwrap.wrap(widgetString, maxChar, break_long_words=True))):
            stringList.append(textwrap.wrap(widgetString, maxChar, break_long_words=True)[x])
        if widgetString is None:
            # print(f"Empty here 2")
            widgetString = "EmptyNode"
        for x in range(len(stringList)):
            if len(stringList[x]) > 1:
                if x == 0:
                    tableMiddle = f"<tr><td ALIGN='LEFT'><FONT POINT-SIZE='11' COLOR='red'><B>{stringList[x][0].capitalize()}</B></FONT>" \
                                  f"<FONT POINT-SIZE='10'>{stringList[x][1:]}</FONT></td></tr>"
                else:
                    tableMiddle += f"<tr><td ALIGN='LEFT'><FONT POINT-SIZE='10'>{stringList[x]}</FONT></td></tr>"
            elif len(stringList[x]) == 0:
                tableMiddle = f"<tr><td ALIGN='LEFT'><FONT POINT-SIZE='10'>{stringList[x].capitalize()}</FONT></td></tr>"
            else:
                tableMiddle = f"<tr><td ALIGN='LEFT'><FONT POINT-SIZE='10'>EmptyNode</FONT></td></tr>"
        return tableMiddle

    def graphHeader(self, mergePdf, label):
        graph = graphviz.Digraph(self.graphName)
        graph.node('A1', shape='none', color='black', label=label)
        graph.render(filename=f"{self.outDir}{self.graphName}", view=False)
        mergePdf.append(PdfFileReader(f"{self.outDir}{self.graphName}.pdf", 'rb'))
        os.remove(graph.render(filename=f"{self.outDir}{self.graphName}", view=False))

    def graphMenu(self, mergePdf):
        graph = graphviz.Digraph(self.graphName)
        graph.attr(shape="circle", color="green", compound="True")

    def graph(self, mergePdf, sheetPage):
        graph = graphviz.Digraph(self.graphName)
        graph.attr(shape="circle", color="green", compound="True")

        keyList = list(self.mainNodeDict.keys())
        for index in range(len(keyList)):
            mainNode = keyList[index]
            try:
                graph.node(self.mainNodeDict.get(mainNode).get('Label'),
                           label=self.mainNodeDict.get(mainNode).get('Widget'),
                           color=self.mainNodeDict.get(mainNode).get('color'),
                           shape=self.mainNodeDict.get(mainNode).get('shape'), fontname='Arial')
            except SyntaxError:
                pass

        keyListEdge = list(self.mainEdgeDict.keys())
        for indexEdge in range(len(keyListEdge)):
            mainEdge = keyListEdge[indexEdge]
            graph.edge(self.mainEdgeDict.get(mainEdge)[0], self.mainEdgeDict.get(mainEdge)[1])
            # print(f"Edge_main: {self.mainEdgeDict.get(mainEdge)[0]}, {self.mainEdgeDict.get(mainEdge)[-1]}")

        # Create subgraph from the main 'graph' tree
        # In order to use the cluster traits, name must start with the string "cluster"
        with graph.subgraph(name="cluster") as subCluster:
            # Compound ensures that the rank and shape of the subgraph would not
            # take precedence over the main graph node structure hierarchy
            subCluster.attr(shape="box", color="red", compound="True")
            # Convert the dictionary used to obtain node values into a list
            keyListSub = list(self.subNodeDict.keys())
            for indexSub in range(len(keyListSub)):
                subNode = keyListSub[indexSub]
                # Create node using the Label and Values derived from the dictionary
                # Dictionary also holds shape and color identity
                try:
                    subCluster.node(self.subNodeDict.get(subNode).get('Label'),
                                    label=self.subNodeDict.get(subNode).get('Widget'),
                                    color=self.subNodeDict.get(subNode).get('color'),
                                    shape=self.subNodeDict.get(subNode).get('shape'), fontname='Arial')
                except SyntaxError:
                    pass
            # Loop over the edge dictionary to make the connections between the nodes
            # Edges will be connecting using the nodes' labels
            # For example, nodes (A0, X) & (A1, Y) will be connected by edge:
            # (A0, A1), which will result in the tree creating the connection that (X -> Y) in the child hierarchy
            keyListEdgeSub = list(self.subEdgeDict.keys())
            for indexEdgeSub in range(len(keyListEdgeSub)):
                subEdge = keyListEdgeSub[indexEdgeSub]
                subCluster.edge(self.subEdgeDict.get(subEdge)[0], self.subEdgeDict.get(subEdge)[1])
                # print(f"Edge_sub: {self.subEdgeDict.get(subEdge)[0]}, {self.subEdgeDict.get(subEdge)[-1]}")

        # print(f'graph {graph}')

        graph.render(filename=f"{self.outDir}{self.graphName}", view=False)
        mergePdf.append(PdfFileReader(f"{self.outDir}{self.graphName}.pdf", 'rb'))

        os.remove(graph.render(filename=f"{self.outDir}{self.graphName}", view=False))

    def graphL(self, mergePdf, sheetPage):
        graph = graphviz.Digraph(self.graphName)
        graph.attr(shape="circle", color="green", compound="True")

        keyList = list(self.mainNodeDictL.keys())
        for index in range(len(keyList)):
            mainNode = keyList[index]
            # print(f"Main node_graph: {mainNode}")
            # print(f"Dict main node: {self.mainNodeDict.get(mainNode)}")
            # print(f"Dict main node: {self.mainNodeDict.get(mainNode).get('Widget')}")
            try:
                graph.node(self.mainNodeDictL.get(mainNode).get('Label'),
                           label=self.mainNodeDictL.get(mainNode).get('Widget'),
                           color=self.mainNodeDictL.get(mainNode).get('color'),
                           shape=self.mainNodeDictL.get(mainNode).get('shape'), fontname='Arial')
            except SyntaxError:
                pass
        keyListEdge = list(self.mainEdgeDictL.keys())
        for indexEdge in range(len(keyListEdge)):
            mainEdge = keyListEdge[indexEdge]
            graph.edge(self.mainEdgeDictL.get(mainEdge)[0], self.mainEdgeDictL.get(mainEdge)[1])

        print(f"Graph: {graph}\n")

        graph.render(filename=f"{self.outDir}{self.graphName}", view=False)
        mergePdf.append(PdfFileReader(f"{self.outDir}{self.graphName}.pdf", 'rb'))
        os.remove(graph.render(filename=f"{self.outDir}{self.graphName}", view=False))


class Spell:

    def __init__(self, fileLocationPath):
        self.sym_spell = SymSpell(max_dictionary_edit_distance=2, prefix_length=7)
        self.sym_spell.load_dictionary("frequency_dictionary_en_82_765.txt", term_index=0, count_index=1)
        self.pathToFile = fileLocationPath
        self.totalTime = 0
        self.tableRows = [["MISSPELLED", "SUGGESTIONS", "SHEETNAME", "ROW:COLUMN", "INDEX", "XPATH"]]

    def spell(self, string, sheetName, cellValueRow, cellValueColumn, sheetIndex, xPath):
        s = time.time()
        # spellChecker = SpellChecker()
        # missSpelled = spellChecker.unknown(spellChecker.split_words(string.encode('utf-8').decode('ascii', 'ignore')))
        string = string.lower()
        missSpelled = string.encode('utf-8').decode('ascii', 'ignore').split()

        # self.sym_spell.
        e = time.time()
        s = time.time()
        # print(e-s)
        for missSpelledWord in missSpelled:
            # Get the one `most likely` answer
            corrections = self.sym_spell.lookup(missSpelledWord, Verbosity.TOP, max_edit_distance=2)
            distance = None
            for correction in corrections:
                if correction.distance == 0:
                    distance = 0
                corrections = str(correction)

            if distance == 0:
                continue

            try:
                corrections = corrections[0:corrections.index(',')]
                temp = [missSpelledWord, str(corrections), sheetName,
                        str(str(cellValueRow) + ":" + str(cellValueColumn)),
                        sheetIndex, xPath]
            except:
                temp = [missSpelledWord, "UNKNOWN", sheetName,
                        str(str(cellValueRow) + ":" + str(cellValueColumn)),
                        sheetIndex, xPath]

            # Get a list of `likely` options
            # print((spell.candidates(missSpelledWord)).)


            self.tableRows.append(temp)
            # print(temp)
        e = time.time()
        self.totalTime = self.totalTime + (e - s)

    def createPDF(self):
        print("Total Checking Time: " + str(self.totalTime))
        s = time.time()
        pdf = PDF()
        pdf.set_top_margin(20)
        pdf.add_page(orientation='L')
        pdf.set_font("Arial", size=9)

        pdf.create_table(table_data=self.tableRows, title="SPELL CHECK", cell_width=[30, 30, 30, 22, 20, 150],
                         align_data='L',
                         align_header='C', x_start='C', data_size=8, title_size=10)

        pdf.ln()
        pdf.output(self.pathToFile)  # os.path.join(self.pathToFile, 'table_class.pdf'))
        e = time.time()
        print("Total Creating Time: " + str(e - s))


sheet = Spell('spellCheck.pdf')


class PDF(FPDF):
    def create_table(self, table_data, title='', data_size=10, title_size=12, align_data='L', align_header='L',
                     cell_width='even', x_start='x_default', emphasize_data=[], emphasize_style=None,
                     emphasize_color=(0, 0, 0)):
        """
        table_data:
                    list of lists with first element being list of headers
        title:
                    (Optional) title of table (optional)
        data_size:
                    the font size of table data
        title_size:
                    the font size fo the title of the table
        align_data:
                    align table data
                    L = left align
                    C = center align
                    R = right align
        align_header:
                    align table data
                    L = left align
                    C = center align
                    R = right align
        cell_width:
                    even: evenly distribute cell/column width
                    uneven: base cell size on lenght of cell/column items
                    int: int value for width of each cell/column
                    list of ints: list equal to number of columns with the widht of each cell / column
        x_start:
                    where the left edge of table should start
        emphasize_data:
                    which data elements are to be emphasized - pass as list
                    emphasize_style: the font style you want emphaized data to take
                    emphasize_color: emphasize color (if other than black)

        """
        default_style = self.font_style
        if emphasize_style == None:
            emphasize_style = default_style

        # default_font = self.font_family
        # default_size = self.font_size_pt
        # default_style = self.font_style
        # default_color = self.color # This does not work

        # Get Width of Columns
        def get_col_widths():
            col_width = cell_width
            if col_width == 'even':
                col_width = self.epw / len(data[
                                               0]) - 1  # distribute content evenly   # epw = effective page width (width of page not including margins)
            elif col_width == 'uneven':
                col_widths = []

                # searching through columns for largest sized cell (not rows but cols)
                for col in range(len(table_data[0])):  # for every row
                    longest = 0
                    for row in range(len(table_data)):
                        cell_value = str(table_data[row][col])
                        value_length = self.get_string_width(cell_value)
                        if value_length > longest:
                            longest = value_length
                    col_widths.append(longest + 4)  # add 4 for padding
                col_width = col_widths
                ### compare columns

            elif isinstance(cell_width, list):
                col_width = cell_width  # TODO: convert all items in list to int
            else:
                # TODO: Add try catch
                col_width = int(col_width)
            return col_width

        # Convert dict to lol
        # Why? because i built it with lol first and added dict func after
        # Is there performance differences?
        if isinstance(table_data, dict):
            header = [key for key in table_data]
            data = []
            for key in table_data:
                value = table_data[key]
                data.append(value)
            # need to zip so data is in correct format (first, second, third --> not first, first, first)
            data = [list(a) for a in zip(*data)]

        else:
            header = table_data[0]
            data = table_data[1:]

        line_height = self.font_size * 2.5

        col_width = get_col_widths()
        # fontSize = title_size + 10
        self.set_font('Arial', "B", 12)
        # self.set_font(style='B')
        self.set_text_color(2, 79, 151)

        # Get starting position of x
        # Determin width of table to get x starting point for centred table
        if x_start == 'C':
            table_width = 0
            if isinstance(col_width, list):
                for width in col_width:
                    table_width += width
            else:  # need to multiply cell width by number of cells to get table width
                table_width = col_width * len(table_data[0])
            # Get x start by subtracting table width from pdf width and divide by 2 (margins)
            margin_width = self.w - table_width
            # TODO: Check if table_width is larger than pdf width

            center_table = margin_width / 2  # only want width of left margin not both
            x_start = center_table
            self.set_x(x_start)
        elif isinstance(x_start, int):
            self.set_x(x_start)
        elif x_start == 'x_default':
            x_start = self.set_x(self.l_margin)

        # TABLE CREATION #

        # # add title
        # if title != '':
        #     self.multi_cell(0, line_height, title, border=0, align='C', split_only=False, ln=3,
        #                     max_line_height=self.font_size,
        #                     markdown=True, print_sh=False)
        #     self.ln(line_height)  # move cursor back to the left margin

        self.set_font(size=data_size)
        self.set_font(style='B')
        self.set_text_color(229, 125, 28)
        # self.set_fill_color(255,255,255)
        # add header
        y1 = self.get_y()
        if x_start:
            x_left = x_start
        else:
            x_left = self.get_x()
        x_right = self.epw + x_left
        if not isinstance(col_width, list):
            if x_start:
                self.set_x(x_start)
            for datum in header:
                self.multi_cell(col_width, line_height, datum, border=0, align=align_header, ln=3,
                                max_line_height=self.font_size)
                x_right = self.get_x()
            self.ln(line_height)  # move cursor back to the left margin
            y2 = self.get_y()
            self.line(x_left, y1, x_right, y1)
            self.line(x_left, y2, x_right, y2)

            for row in data:
                if x_start:  # not sure if I need this
                    self.set_x(x_start)
                for datum in row:
                    if datum in emphasize_data:
                        self.set_text_color(*emphasize_color)
                        self.set_font(style=emphasize_style)
                        self.multi_cell(col_width, line_height, datum, border=0, align=align_data, ln=3,
                                        max_line_height=self.font_size)
                        self.set_text_color(0, 0, 0)
                        self.set_font(style=default_style)
                    else:
                        self.multi_cell(col_width, line_height, datum, border=0, align=align_data, ln=3,
                                        max_line_height=self.font_size)  # ln = 3 - move cursor to right with same vertical offset # this uses an object named self
                self.ln(line_height)  # move cursor back to the left margin

        else:
            if x_start:
                self.set_x(x_start)
            for i in range(len(header)):
                datum = header[i]
                self.multi_cell(col_width[i], line_height, datum, border=0, align=align_header, ln=3,
                                max_line_height=self.font_size)
                x_right = self.get_x()
            self.ln(line_height)  # move cursor back to the left margin
            y2 = self.get_y()
            self.line(x_left, y1, x_right, y1)
            self.line(x_left, y2, x_right, y2)

            self.set_font()
            self.set_text_color(0, 0, 0)

            for i in range(len(data)):
                if x_start:
                    self.set_x(x_start)
                row = data[i]
                for i in range(len(row)):
                    datum = row[i]
                    if not isinstance(datum, str):
                        datum = str(datum)
                    adjusted_col_width = col_width[i]
                    if datum in emphasize_data:
                        self.set_text_color(*emphasize_color)
                        self.set_font(style=emphasize_style)
                        self.multi_cell(adjusted_col_width, line_height, datum, border=0, align=align_data, ln=3,
                                        max_line_height=self.font_size)
                        self.set_text_color(0, 0, 0)
                        self.set_font(style=default_style)
                    else:
                        self.multi_cell(adjusted_col_width, line_height, datum, border=0, align=align_data, ln=3,
                                        max_line_height=self.font_size)  # ln = 3 - move cursor to right with same vertical offset # this uses an object named self
                self.ln(line_height)  # move cursor back to the left margin
        y3 = self.get_y()
        # self.line(x_left, y3, x_right, y3)


class header_and_footer:

    def __init__(self, input_pdf, overwrite_pdf, spellCheck_pdf):
        self.input_pdf = input_pdf
        self.overwrite_pdf = overwrite_pdf
        self.spellCheck_pdf = spellCheck_pdf

    def create_header(self, page_width, page_height):
        scale = 3
        header = Image.new(mode="RGB", size=(round(page_width * scale), round(75 * scale)),
                           color=(255, 255, 255))  # color=(37, 166, 218))
        # color=(37, 166, 218)) - color=(255,255,255))#
        logo = Image.open(self.overwrite_pdf.get("full_logo_path"))
        logo.thumbnail((100 * scale, 100 * scale))

        # Left Side
        left_side_text = ("Sheet Name: \n> " + split_string(self.overwrite_pdf.get("sheet_info")[0]) +
                          "Pre Requisites: \n> " + split_string(self.overwrite_pdf.get("pre_requisites")[0]) +
                          split_string("Step Number: " + self.overwrite_pdf.get("step_number")[0]))

        # Center
        header.paste(logo, (round(header.width / 2 - logo.width / 2), round(header.height / 2 - logo.height / 2)))

        # Right Side
        right_side_text = (split_string("Version Number: " + self.overwrite_pdf.get("version_number")) +
                           split_string("Server: " + self.overwrite_pdf.get("server")))

        # Font Setup
        font_size = 8 * scale
        font = ImageFont.truetype("arial.ttf", font_size)
        # rightW, rightH = font.getsize(right_side_text)
        leftH = header.height
        leftW = header.width

        while leftH > (header.height - 5) / scale:
            font_size -= 1
            font = ImageFont.truetype("arial.ttf", font_size)
            leftW, leftH = font.getsize(left_side_text)

        d = ImageDraw.Draw(header)

        d.text((10 * scale, round((header.height / 2) - (leftH * scale))), left_side_text, font=font, fill=(0, 0, 0))
        d.text((header.width - 38 * 2.5 * scale, round((header.height / 2) - (leftH * scale))), right_side_text,
               font=font, fill=(0, 0, 0))

        header.save("header.png")
        return header.width / scale, header.height / scale

    def create_footer(self, page_width, page_height):
        scale = 3
        footer = Image.new(mode="RGB", size=(round(page_width * scale), round(75 * scale)),
                           color=(255, 255, 255))  # color=(37, 166, 218))
        # color=(37, 166, 218)) - color=(255,255,255))#
        logo = Image.open(self.overwrite_pdf.get("small_logo_path"))
        logo.thumbnail((50 * scale, 50 * scale))

        # Left Side
        left_side_text = ("Created On: " + self.overwrite_pdf.get("created_on") +
                          ("\nPage Number: " + self.overwrite_pdf.get("page_number")))

        # Right Side
        footer.paste(logo, (round(footer.width - logo.width - 10 * scale), round(footer.height / 2 - logo.height / 2)))

        # Font Setup
        font_size = 10
        font = ImageFont.truetype("arial.ttf", font_size * scale)
        leftH = footer.height
        leftW = footer.width

        while leftH > (footer.height - 5) / scale:
            font_size -= 1
            font = ImageFont.truetype("arial.ttf", font_size * scale)
            leftW, leftH = font.getsize(left_side_text)

        d = ImageDraw.Draw(footer)

        d.text((10 * scale, round((footer.height / 2) - (5 * scale))), left_side_text, font=font, fill=(0, 0, 0))

        footer.save("footer.png")
        return footer.width / scale, footer.height / scale

    def create_spell_header(self, page_width, page_height):
        scale = 3
        header = Image.new(mode="RGB", size=(round(page_width * scale), round(50 * scale)),
                           color=(255, 255, 255))  # color=(37, 166, 218))
        # color=(37, 166, 218)) - color=(255,255,255))#
        logo = Image.open(self.overwrite_pdf.get("full_logo_path"))
        logo.thumbnail((100 * scale, 100 * scale))

        # Center
        header.paste(logo, (40, round(header.height / 2 - logo.height / 2) + 5))

        # Right Side
        right_side_text = ("Version Number: " + split_string(self.overwrite_pdf.get("version_number")))

        # Font Setup
        font_size = 8 * scale
        font = ImageFont.truetype("arial.ttf", font_size)
        # rightW, rightH = font.getsize(right_side_text)
        leftH = header.height
        leftW = header.width

        while leftH > (header.height - 5) / scale:
            font_size -= 1
            font = ImageFont.truetype("arial.ttf", font_size)
            leftW, leftH = font.getsize(right_side_text)

        d = ImageDraw.Draw(header)

        d.text((header.width - 38 * 2.5 * scale, round((header.height / 2) - (leftH * scale))), right_side_text,
               font=font, fill=(0, 0, 0))

        header.save("header.png")
        return header.width / scale, header.height / scale

    def create_spell_footer(self, page_width, page_height, page_counter):
        scale = 3
        footer = Image.new(mode="RGB", size=(round(page_width * scale), round(50 * scale)),
                           color=(255, 255, 255))  # color=(37, 166, 218))
        # color=(37, 166, 218)) - color=(255,255,255))#

        logo = Image.open(self.overwrite_pdf.get("small_logo_path"))
        logo.thumbnail((40 * scale, 40 * scale))

        # Left Side
        left_side_text = ("Created On: " + self.overwrite_pdf.get("created_on") +
                          ("\nPage Number: " + str(page_counter)))

        # Right Side
        footer.paste(logo, (
            round(footer.width - logo.width - 10 * scale - 20), round(footer.height / 2 - logo.height / 2 + 2)))

        # Font Setup
        font_size = 10
        font = ImageFont.truetype("arial.ttf", font_size * scale)
        leftH = footer.height
        leftW = footer.width

        while leftH > (footer.height - 5) / scale:
            font_size -= 1
            font = ImageFont.truetype("arial.ttf", font_size * scale)
            leftW, leftH = font.getsize(left_side_text)

        w, h = 220, 190
        shape = [(40, 0), (2470, 0)]

        d = ImageDraw.Draw(footer)
        d.text((10 * scale + 10, round((footer.height / 2) - (5 * scale))), left_side_text, font=font, fill=(0, 0, 0))
        d.line(shape, width=3, fill="#000000")

        footer.save("footer.png")
        return footer.width / scale, footer.height / scale

    def draw_header_and_footer(self):

        # ----------- Add Margin By Resize ------------

        start_time = time.time()

        docLocation = open(self.input_pdf, 'rb')
        doc = PdfFileReader(docLocation)

        newDoc = PdfFileWriter()
        for pageCounter in range(0, doc.getNumPages()):
            originalPage = doc.getPage(pageCounter)
            newPage = PageObject.createBlankPage(None, originalPage.mediaBox.getWidth(),
                                                 originalPage.mediaBox.getHeight() + 100)
            newPage.mergeScaledTranslatedPage(originalPage, 1, 0, 50)
            newDoc.addPage(newPage)
        output = open('Resize.pdf', 'wb')
        newDoc.write(output)
        docLocation.close()
        output.close()
        os.remove(self.input_pdf)
        os.rename('Resize.pdf', self.input_pdf)
        # shutil.copyfile('Resize.pdf', self.input_pdf)
        # os.remove('Resize.pdf')
        end_time = time.time()

        print("Total Margin Resize Time: " + str(end_time - start_time))

        # ----------- Add Header And Footer ------------

        start_time = time.time()

        full_logo_file = self.overwrite_pdf.get("full_logo_path")
        small_logo_file = self.overwrite_pdf.get("small_logo_path")
        doc = fitz.open(self.input_pdf)

        page_counter = 0
        for page in doc:
            page_width = round(doc[page_counter].rect.width)
            page_height = round(doc[page_counter].rect.height)

            headerW, headerH = header_and_footer.create_header(self, page_width, page_height)
            footerW, footerH = header_and_footer.create_footer(self, page_width, page_height)
            header_rect = fitz.Rect(0, 0, page_width, headerH + 10)
            footer_rect = fitz.Rect(0, page_height - footerH - 10, page_width, page_height)
            page.insert_image(header_rect, filename="header.png")
            page.insert_image(footer_rect, filename="footer.png")
            page_counter += 1
            self.overwrite_pdf["page_number"] = str(page_counter + 1)
            self.overwrite_pdf["sheet_info"].pop(0)
            self.overwrite_pdf["pre_requisites"].pop(0)
            self.overwrite_pdf["step_number"].pop(0)

        end_time = time.time()

        print("Total Add Header And Footer Time: " + str(end_time - start_time))
        # ----------- Add Spell Check ------------

        start_time = time.time()

        totalPages = page_counter + 1
        page_counter = 0
        doc.saveIncr()
        doc.close()
        spellDoc = fitz.open(self.spellCheck_pdf)
        page_width = round(spellDoc[0].rect.width)
        page_height = round(spellDoc[0].rect.height)
        headerW, headerH = header_and_footer.create_spell_header(self, page_width, page_height)
        for page in spellDoc:
            footerW, footerH = header_and_footer.create_spell_footer(self, page_width, page_height, totalPages)
            header_rect = fitz.Rect(0, 0, page_width, headerH + 10)
            footer_rect = fitz.Rect(0, page_height - footerH - 10, page_width, page_height)
            page.insert_image(header_rect, filename="header.png")
            page.insert_image(footer_rect, filename="footer.png")
            page_counter += 1
            totalPages += 1

        spellDoc.saveIncr()
        spellDoc.close()
        os.remove("header.png")
        os.remove("footer.png")

        shutil.copyfile(self.input_pdf, "temp.pdf")
        os.remove(self.input_pdf)

        merger = PdfFileMerger()

        merger.append("temp.pdf")
        merger.append(self.spellCheck_pdf)

        merger.write(self.input_pdf)
        merger.close()
        os.remove("temp.pdf")
        os.remove(self.spellCheck_pdf)

        end_time = time.time()

        print("Total Spell Check Header and Footer Time: " + str(end_time - start_time))


def split_string(text):
    n = 32
    finalWrap = ''
    for x in range(len(textwrap.wrap(text, n, break_long_words=False))):
        finalWrap += textwrap.wrap(text, n, break_long_words=False)[x] + '\n'
    return finalWrap


# ---------------- What you need to use -----------------


def add_header_and_footer(input_file, sheet_info, pre_requisites, step_number, full_logo_path, small_logo_path,
                          version_number,
                          server, created_on, spellCheck_pdf):
    info_pdf = {
        "sheet_info": sheet_info,
        "pre_requisites": pre_requisites,
        "step_number": step_number,
        "full_logo_path": full_logo_path,
        "small_logo_path": small_logo_path,
        "version_number": version_number,
        "server": server,
        "created_on": created_on,
        "page_number": "1",
    }

    operation = header_and_footer(input_file, info_pdf, spellCheck_pdf)
    operation.draw_header_and_footer()


def main(excelDir, graphName, outDir, serverURL, labelNum, maxChar, logoLink):
    wb = GraphNodes(excelDir, graphName, outDir, serverURL, labelNum, maxChar, logoLink)
    wb.sort()
    # print(f"Path: {pathlib.Path().resolve()}")


if __name__ == "__main__":
    excelPath = sys.argv[1]
    graphvizName = sys.argv[2]
    outputDirectory = sys.argv[3]
    serverUrl = sys.argv[4]
    labelType = int(sys.argv[5])
    logoLink = sys.argv[6]
    # excelPath = "C:/Users/Shaya/PycharmProjects/AiTestPro-TestflowGraph/pyinstaller/pythonExe/UpdatedTest_DGErrorLeema.xlsx"
    # excelPath = "C:/Users/Shaya/PycharmProjects/AiTestPro-TestflowGraph/pyinstaller/pythonExe/RogersRakeshMarch9.xlsx"
    # C:/Users/Shaya/PycharmProjects/AiTestPro-TestflowGraph/pyinstaller/pythonExe/rakeshNewDGFeb23_v2.xlsx
    # excelPath = "RogersRakeshMarch9.xlsx"
    # graphvizName = "RogersRakeshMarch9"
    # outputDirectory = ""
    # serverUrl = "url"
    # labelType = 2
    # logoLink = 'Default'
    # logoLink = 'https://images.ctfassets.net/8utyj17y1gom/2JXnJrw79KkQOiOMia8Au2/9cbfc2fda65fa5a2127d5b0668d84765/rogers-brand.png'
    maxCharLimit = 20
    main(excelPath, graphvizName, outputDirectory, serverUrl, labelType, maxCharLimit, logoLink)

e = time.time()

print("totalTime: " + str(e - s))